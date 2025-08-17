import os
import json
import random
import string
import logging
import unicodedata
import io
from datetime import datetime, timedelta
from functools import wraps

# Imports pour les exports PDF et Excel
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab n'est pas installé. L'export PDF ne sera pas disponible.")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl n'est pas installé. L'export Excel ne sera pas disponible.")

def get_blank_location(global_blank_index, sentences):
    """Détermine à quelle phrase et à quel indice dans cette phrase correspond un indice global de blanc"""
    blank_count = 0
    for idx, sentence in enumerate(sentences):
        blanks_in_sentence = sentence.count('___')
        if blank_count <= global_blank_index < blank_count + blanks_in_sentence:
            # Calculer l'indice local du blanc dans cette phrase
            local_blank_index = global_blank_index - blank_count
            return idx, local_blank_index
        blank_count += blanks_in_sentence
    return -1, -1

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session, current_app, abort
from flask_login import login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from flask_wtf import FlaskForm, CSRFProtect
from wtforms import StringField, TextAreaField, MultipleFileField
from wtforms.validators import DataRequired
from flask_wtf.csrf import generate_csrf

from extensions import db, init_extensions, login_manager
from models import User, Class, Course, Exercise, ExerciseAttempt, CourseFile, student_class_association, course_exercise
from payment_routes import payment_bp
from forms import ExerciseForm
from modified_submit import bp as exercise_bp





logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('flask_app.log')
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.logger.setLevel(logging.DEBUG)


# Route de diagnostic pour tous les exercices fill_in_blanks
@app.route('/debug-all-fill-in-blanks')
def debug_all_fill_in_blanks():
    # Route de diagnostic pour analyser tous les exercices fill_in_blanks
    if not current_user.is_authenticated or not current_user.is_admin:
        return "Accès non autorisé", 403
        
    results = []
    
    # En-tête
    results.append("<h1>DIAGNOSTIC TOUS LES EXERCICES FILL_IN_BLANKS</h1>")
    
    # 1. Environnement
    results.append("<h2>1. ENVIRONNEMENT</h2>")
    
    # Vérifier les variables d'environnement
    env_vars = {
        'FLASK_ENV': os.environ.get('FLASK_ENV', 'non défini'),
        'DATABASE_URL': os.environ.get('DATABASE_URL', 'non défini')[:10] + '...' if os.environ.get('DATABASE_URL') else 'non défini',
        'RAILWAY_ENVIRONMENT': os.environ.get('RAILWAY_ENVIRONMENT', 'non défini'),
        'PORT': os.environ.get('PORT', 'non défini')
    }
    
    results.append("<h3>Variables d'environnement:</h3>")
    for key, value in env_vars.items():
        results.append(f"<p>{key}: {value}</p>")
    
    # 2. Liste des exercices
    results.append("<h2>2. LISTE DES EXERCICES FILL_IN_BLANKS</h2>")
    
    try:
        exercises = Exercise.query.filter_by(exercise_type='fill_in_blanks').all()
        results.append(f"<p>Nombre d'exercices fill_in_blanks: {len(exercises)}</p>")
        
        if exercises:
            results.append("<table border='1' style='border-collapse: collapse; width: 100%;'>")
            results.append("<tr><th>ID</th><th>Titre</th><th>Image</th><th>Blancs</th><th>Mots</th><th>Cohérence</th></tr>")
            
            for ex in exercises:
                try:
                    content = json.loads(ex.content)
                    
                    # Compter les blancs
                    total_blanks = 0
                    
                    if 'sentences' in content:
                        sentences_blanks = sum(s.count('___') for s in content['sentences'])
                        total_blanks = sentences_blanks
                    elif 'text' in content:
                        text_blanks = content['text'].count('___')
                        total_blanks = text_blanks
                    
                    # Compter les mots
                    words = []
                    if 'words' in content:
                        words = content['words']
                    elif 'available_words' in content:
                        words = content['available_words']
                    
                    # Vérifier la cohérence
                    coherence = "✓" if total_blanks == len(words) else "✗"
                    coherence_color = "green" if total_blanks == len(words) else "red"
                    
                    # Image
                    has_image = "✓" if ex.image_path else "✗"
                    image_color = "green" if ex.image_path else "gray"
                    
                    results.append(f"<tr>")
                    results.append(f"<td>{ex.id}</td>")
                    results.append(f"<td>{ex.title}</td>")
                    results.append(f"<td style='color: {image_color};'>{has_image}</td>")
                    results.append(f"<td>{total_blanks}</td>")
                    results.append(f"<td>{len(words)}</td>")
                    results.append(f"<td style='color: {coherence_color};'>{coherence}</td>")
                    results.append(f"</tr>")
                except Exception as e:
                    results.append(f"<tr><td>{ex.id}</td><td>{ex.title}</td><td colspan='4' style='color: red;'>Erreur: {str(e)}</td></tr>")
            
            results.append("</table>")
        else:
            results.append("<p>Aucun exercice fill_in_blanks trouvé.</p>")
    except Exception as e:
        results.append(f"<p style='color: red;'>Erreur lors de la récupération des exercices: {str(e)}</p>")
    
    # 3. Liste des exercices word_placement
    results.append("<h2>3. LISTE DES EXERCICES WORD_PLACEMENT</h2>")
    
    try:
        exercises = Exercise.query.filter_by(exercise_type='word_placement').all()
        results.append(f"<p>Nombre d'exercices word_placement: {len(exercises)}</p>")
        
        if exercises:
            results.append("<table border='1' style='border-collapse: collapse; width: 100%;'>")
            results.append("<tr><th>ID</th><th>Titre</th><th>Image</th><th>Blancs</th><th>Mots</th><th>Cohérence</th></tr>")
            
            for ex in exercises:
                try:
                    content = json.loads(ex.content)
                    
                    # Compter les blancs
                    total_blanks = 0
                    
                    if 'sentences' in content:
                        sentences_blanks = sum(s.count('___') for s in content['sentences'])
                        total_blanks = sentences_blanks
                    elif 'text' in content:
                        text_blanks = content['text'].count('___')
                        total_blanks = text_blanks
                    
                    # Compter les mots
                    words = []
                    if 'words' in content:
                        words = content['words']
                    elif 'available_words' in content:
                        words = content['available_words']
                    
                    # Vérifier la cohérence
                    coherence = "✓" if total_blanks == len(words) else "✗"
                    coherence_color = "green" if total_blanks == len(words) else "red"
                    
                    # Image
                    has_image = "✓" if ex.image_path else "✗"
                    image_color = "green" if ex.image_path else "gray"
                    
                    results.append(f"<tr>")
                    results.append(f"<td>{ex.id}</td>")
                    results.append(f"<td>{ex.title}</td>")
                    results.append(f"<td style='color: {image_color};'>{has_image}</td>")
                    results.append(f"<td>{total_blanks}</td>")
                    results.append(f"<td>{len(words)}</td>")
                    results.append(f"<td style='color: {coherence_color};'>{coherence}</td>")
                    results.append(f"</tr>")
                except Exception as e:
                    results.append(f"<tr><td>{ex.id}</td><td>{ex.title}</td><td colspan='4' style='color: red;'>Erreur: {str(e)}</td></tr>")
            
            results.append("</table>")
        else:
            results.append("<p>Aucun exercice word_placement trouvé.</p>")
    except Exception as e:
        results.append(f"<p style='color: red;'>Erreur lors de la récupération des exercices: {str(e)}</p>")
    
    # 4. Test de la logique de scoring
    results.append("<h2>4. TEST DE LA LOGIQUE DE SCORING</h2>")
    
    # Test avec sentences
    results.append("<h3>Test avec sentences</h3>")
    test_content_sentences = {
        "sentences": ["Le ___ mange une ___ rouge."],
        "words": ["chat", "pomme"]
    }
    
    # Simuler des réponses utilisateur parfaites
    user_answers_sentences = {
        'answer_0': 'chat',
        'answer_1': 'pomme'
    }
    
    # Calculer le score
    try:
        total_blanks = sum(s.count('___') for s in test_content_sentences['sentences'])
        correct_blanks = 0
        
        for i in range(total_blanks):
            answer_key = f'answer_{i}'
            user_answer = user_answers_sentences.get(answer_key, '')
            correct_answer = test_content_sentences['words'][i] if i < len(test_content_sentences['words']) else ''
            
            if user_answer.lower() == correct_answer.lower():
                correct_blanks += 1
        
        score = round((correct_blanks / total_blanks) * 100) if total_blanks > 0 else 0
        
        results.append(f"<p>Score avec sentences: {correct_blanks}/{total_blanks} = {score}%</p>")
        if score == 100:
            results.append("<p style='color: green;'>✓ Test sentences réussi!</p>")
        else:
            results.append("<p style='color: red;'>✗ Test sentences échoué!</p>")
    except Exception as e:
        results.append(f"<p style='color: red;'>Erreur test sentences: {str(e)}</p>")
    
    # Test avec text
    results.append("<h3>Test avec text</h3>")
    test_content_text = {
        "text": "Le ___ mange une ___ rouge.",
        "words": ["chat", "pomme"]
    }
    
    # Simuler des réponses utilisateur parfaites
    user_answers_text = {
        'answer_0': 'chat',
        'answer_1': 'pomme'
    }
    
    # Calculer le score
    try:
        total_blanks = test_content_text['text'].count('___')
        correct_blanks = 0
        
        for i in range(total_blanks):
            answer_key = f'answer_{i}'
            user_answer = user_answers_text.get(answer_key, '')
            correct_answer = test_content_text['words'][i] if i < len(test_content_text['words']) else ''
            
            if user_answer.lower() == correct_answer.lower():
                correct_blanks += 1
        
        score = round((correct_blanks / total_blanks) * 100) if total_blanks > 0 else 0
        
        results.append(f"<p>Score avec text: {correct_blanks}/{total_blanks} = {score}%</p>")
        if score == 100:
            results.append("<p style='color: green;'>✓ Test text réussi!</p>")
        else:
            results.append("<p style='color: red;'>✗ Test text échoué!</p>")
    except Exception as e:
        results.append(f"<p style='color: red;'>Erreur test text: {str(e)}</p>")
    
    # 5. Conclusion
    results.append("<h2>5. CONCLUSION</h2>")
    results.append("<p>Si tous les tests ci-dessus sont réussis (affichés en vert), la logique de scoring est correcte.</p>")
    results.append("<p>Vérifiez particulièrement:</p>")
    results.append("<ul>")
    results.append("<li>Que le nombre de blancs correspond au nombre de mots pour chaque exercice</li>")
    results.append("<li>Que les tests de scoring donnent bien 100%</li>")
    results.append("<li>Que les exercices problématiques sont identifiés (marqués en rouge)</li>")
    results.append("</ul>")
    
    return "<br>".join(results)
# Configuration selon l'environnement
config_name = os.environ.get('FLASK_ENV', 'development')
if config_name == 'production':
    from config import ProductionConfig
    app.config.from_object(ProductionConfig)
else:
    app.config['SECRET_KEY'] = 'your-secret-key-here'
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///app.db'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Global request logging removed to restore normal operation

# Initialize CSRF protection
csrf = CSRFProtect()
csrf.init_app(app)

# Register blueprints
app.register_blueprint(exercise_bp, url_prefix='/exercise')

# Import and register payment blueprint
from payment_routes import payment_bp
app.register_blueprint(payment_bp)

# Ajout du filtre shuffle pour Jinja2
@app.template_filter('shuffle')
def shuffle_filter(seq):
    try:
        result = list(seq)
        random.shuffle(result)
        return result
    except:
        return seq

# Ajout du filtre chr pour Jinja2 (conversion nombre -> caractère ASCII)
@app.template_filter('chr')
def chr_filter_registered(value):
    """Convertit un nombre en caractère ASCII (65 -> 'A', 66 -> 'B', etc.)"""
    try:
        return chr(int(value))
    except (ValueError, TypeError):
        return str(value)

app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max-limit

# S'assurer que le dossier d'upload existe
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Configuration de l'extension pour les fichiers
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Initialisation des extensions
from extensions import init_extensions
init_extensions(app)

# Initialisation automatique de la base de données
def init_database():
    """Initialise la base de données de manière sécurisée"""
    try:
        with app.app_context():
            # Créer toutes les tables si elles n'existent pas
            db.create_all()
            app.logger.info("Tables de base de donnees creees avec succes")
        
        # Créer le compte administrateur par défaut si nécessaire
        admin_email = os.environ.get('ADMIN_EMAIL', 'admin@classesnumeriques.com')
        admin_user = User.query.filter_by(email=admin_email).first()
        
        if not admin_user:
            from werkzeug.security import generate_password_hash
            admin_user = User(
                username='admin',
                email=admin_email,
                name='Administrateur',
                password_hash=generate_password_hash('AdminSecure2024!'),
                role='admin',
                subscription_status='approved',
                subscription_type='admin',
                approved_by='system'
            )
            db.session.add(admin_user)
            db.session.commit()
            app.logger.info(f"Compte administrateur cree: {admin_email}")
        else:
            app.logger.info(f"Compte administrateur existant: {admin_email}")
        
            # Approuver automatiquement mr.zahiri@gmail.com et lui donner les droits admin
            zahiri_user = User.query.filter_by(email='mr.zahiri@gmail.com').first()
            if zahiri_user:
                zahiri_user.subscription_status = 'approved'
                zahiri_user.role = 'admin'  # Donner les droits admin
                zahiri_user.subscription_type = 'admin'
                zahiri_user.approved_by = 'system'
                db.session.commit()
                app.logger.info("mr.zahiri@gmail.com approuve et promu administrateur")
            else:
                app.logger.info("Compte mr.zahiri@gmail.com non trouve - sera approuve a la creation")
                
    except Exception as e:
        app.logger.error(f"Erreur lors de l'initialisation de la base: {e}")
        # Ne pas faire planter l'app, continuer quand même

# Initialiser la base de données seulement si on n'est pas en mode import
if __name__ == '__main__' or os.environ.get('FLASK_ENV') == 'production':
    init_database()

# Enregistrement des blueprints (déjà fait ligne 48)

# Fonctions pour les filtres Jinja2
def enumerate_filter(iterable, start=0):
    return enumerate(iterable, start=start)

def from_json_filter(value):
    if value is None:
        return None
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return value

def chr_filter(value):
    """Convertit un nombre en caractère ASCII (65 -> 'A', 66 -> 'B', etc.)"""
    try:
        return chr(int(value))
    except (ValueError, TypeError):
        return str(value)

def tojson_filter(value, indent=None):
    return json.dumps(value, indent=indent, ensure_ascii=False)

def get_file_icon(filename):
    """Retourne l'icône Font Awesome appropriée en fonction de l'extension du fichier"""
    extension = filename.lower().split('.')[-1] if '.' in filename else ''
    
    icon_mapping = {
        'pdf': 'fa-file-pdf',
        'doc': 'fa-file-word',
        'docx': 'fa-file-word',
        'xls': 'fa-file-excel',
        'xlsx': 'fa-file-excel',
        'ppt': 'fa-file-powerpoint',
        'pptx': 'fa-file-powerpoint',
        'txt': 'fa-file-alt',
        'jpg': 'fa-file-image',
        'jpeg': 'fa-file-image',
        'png': 'fa-file-image',
        'gif': 'fa-file-image',
        'zip': 'fa-file-archive',
        'rar': 'fa-file-archive',
        '7z': 'fa-file-archive',
    }
    
    return icon_mapping.get(extension, 'fa-file')  # fa-file est l'icône par défaut

app.jinja_env.globals.update(get_file_icon=get_file_icon)

# Enregistrement des filtres Jinja2
app.jinja_env.filters['enumerate'] = enumerate_filter
app.jinja_env.filters['from_json'] = from_json_filter
app.jinja_env.filters['tojson'] = tojson_filter

# Décorateurs
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Accès réservé aux administrateurs.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_teacher:
            flash('Accès réservé aux enseignants.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def sanitize_filename(filename):
    # Supprimer les accents
    filename = ''.join(c for c in unicodedata.normalize('NFD', filename)
                      if unicodedata.category(c) != 'Mn')
    # Remplacer les espaces par des underscores
    filename = filename.replace(' ', '_')
    # Garder uniquement les caractères alphanumériques et quelques caractères spéciaux
    filename = ''.join(c for c in filename if c.isalnum() or c in '._-')
    return filename

def generate_unique_filename(original_filename):
    # Séparer le nom de fichier et l'extension
    name, ext = os.path.splitext(original_filename)
    # Générer un timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    # Générer une chaîne aléatoire
    random_string = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
    # Combiner le tout
    return f"{name}_{timestamp}_{random_string}{ext}"

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@app.after_request
def add_csrf_token(response):
    if 'csrf_token' not in request.cookies:
        response.set_cookie('csrf_token', generate_csrf())
    return response

@app.before_request
def log_request_info():
    logger.debug('Headers: %s', request.headers)
    logger.debug('Body: %s', request.get_data())
    logger.debug('URL: %s', request.url)
    logger.debug('Method: %s', request.method)
    logger.debug('Path: %s', request.path)

# Route pour servir les fichiers uploadés
@app.route('/static/uploads/<path:filename>')
def uploaded_file(filename):
    """Sert les fichiers uploadés depuis le dossier static/uploads"""
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except FileNotFoundError:
        logger.error(f"Fichier non trouvé: {filename}")
        return "Fichier non trouvé", 404

# Routes
@app.route('/')
def index():
    try:
        if current_user.is_authenticated:
            if current_user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif current_user.is_teacher:
                return redirect(url_for('teacher_dashboard'))
            else:  # student
                return redirect(url_for('view_student_classes'))
        else:
            # Utilisateurs non connectés voient la belle page d'accueil moderne avec bouton S'abonner
            return render_template('login.html')
    except Exception as e:
        # En cas d'erreur avec current_user, afficher la page d'accueil par défaut
        app.logger.error(f"Erreur route index: {str(e)}")
        return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember_me = request.form.get('remember_me') == '1'
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            # Vérifier le statut d'abonnement avant la connexion
            if user.subscription_status == 'suspended':
                flash('Votre compte a été suspendu. Contactez l\'administrateur pour plus d\'informations.', 'error')
                return render_template('login.html')
            elif user.subscription_status == 'rejected':
                flash('Votre demande d\'abonnement a été rejetée. Contactez l\'administrateur.', 'error')
                return render_template('login.html')
            elif user.subscription_status == 'pending':
                # Exception spéciale pour mr.zahiri@gmail.com - accès admin direct
                if user.email == 'mr.zahiri@gmail.com':
                    try:
                        user.subscription_status = 'approved'
                        user.role = 'teacher'  # Changé en teacher pour avoir accès à la création d'exercices
                        user.subscription_type = 'admin'
                        user.approved_by = None  # Pas d'ID admin spécifique pour l'auto-approbation
                        user.approval_date = datetime.utcnow()
                        db.session.commit()
                        app.logger.info("✅ mr.zahiri@gmail.com auto-approuvé en tant qu'enseignant-admin")
                    except Exception as e:
                        app.logger.error(f"Erreur lors de l'auto-approbation: {e}")
                        # Continuer quand même la connexion
                        pass
                else:
                    # Permettre la connexion même en attente pour qu'ils puissent voir le bouton "S'abonner"
                    flash('Votre compte est en attente de validation. Vous pouvez effectuer un paiement pour accélérer le processus.', 'warning')
            elif user.subscription_status == 'paid' and user.role != 'admin':
                flash('Votre paiement a été reçu. En attente de validation par l\'administrateur.', 'info')
                return render_template('login.html')
            
            # Connexion autorisée pour les comptes approuvés ou les administrateurs
            login_user(user, remember=remember_me)
            flash('Connexion réussie !', 'success')
            
            # Redirection intelligente selon le rôle
            # Exception spéciale pour mr.zahiri@gmail.com - toujours vers teacher_dashboard
            if user.email == 'mr.zahiri@gmail.com':
                return redirect(url_for('teacher_dashboard'))
            elif user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:  # student
                return redirect(url_for('view_student_classes'))
        else:
            flash('Email ou mot de passe incorrect.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Vous avez été déconnecté.', 'info')
    return redirect(url_for('index'))

@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    # Récupérer les utilisateurs en attente d'approbation
    pending_users = User.query.filter_by(subscription_status='pending').all()
    paid_users = User.query.filter_by(subscription_status='paid').all()
    
    # Récupérer les statistiques générales
    total_users = User.query.count()
    total_teachers = User.query.filter_by(role='teacher').count()
    total_students = User.query.filter_by(role='student').count()
    total_exercises = Exercise.query.count()
    
    return render_template('admin/dashboard.html', 
                           pending_users=pending_users,
                           paid_users=paid_users,
                           total_users=total_users,
                           total_teachers=total_teachers,
                           total_students=total_students,
                           total_exercises=total_exercises)

@app.route('/subscription-choice')
def subscription_choice():
    """Page de choix d'abonnement pour les utilisateurs non connectés"""
    return render_template('subscription_choice.html')

@app.route('/register/teacher', methods=['GET', 'POST'])
def register_teacher():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        school_name = request.form.get('school_name')
        
        if User.query.filter_by(email=email).first():
            flash('Cet email est déjà utilisé.', 'error')
            return redirect(url_for('register_teacher'))
        
        # Générer un username à partir de l'email
        username = email.split('@')[0]
        
        user = User(
            username=username,
            name=name,
            email=email,
            role='teacher',
            school_name=school_name
        )
        user.set_password(password)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash('Inscription réussie ! Vous pouvez maintenant vous connecter.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash('Erreur lors de l\'inscription.', 'error')
            print(f"Erreur d'inscription : {str(e)}")
    
    return render_template('register_teacher.html')

@app.route('/register/school', methods=['GET', 'POST'])
def register_school():
    """Route d'inscription pour les écoles (directeurs/responsables)"""
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        school_name = request.form.get('school_name')
        
        if User.query.filter_by(email=email).first():
            flash('Cet email est déjà utilisé.', 'error')
            return redirect(url_for('register_school'))
        
        # Générer un username à partir de l'email
        username = email.split('@')[0]
        
        user = User(
            username=username,
            name=name,
            email=email,
            role='teacher',  # Les responsables d'école sont aussi des enseignants
            school_name=school_name,
            subscription_type='school'  # Type d'abonnement école
        )
        user.set_password(password)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash('Inscription école réussie ! Vous pouvez maintenant vous connecter et procéder au paiement de 80€.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash('Erreur lors de l\'inscription.', 'error')
            app.logger.error(f"Erreur d'inscription école : {str(e)}")
    
    return render_template('register.html', user_type='school')

@app.route('/register/student', methods=['GET', 'POST'])
def register_student():
    if request.method == 'POST':
        app.logger.info("Données du formulaire reçues : %s", request.form)
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        app.logger.info("Valeurs extraites - username: %s, email: %s", username, email)
        
        # Vérifier que tous les champs sont remplis
        if not all([username, email, password, confirm_password]):
            missing_fields = []
            if not username: missing_fields.append('nom d\'utilisateur')
            if not email: missing_fields.append('email')
            if not password: missing_fields.append('mot de passe')
            if not confirm_password: missing_fields.append('confirmation du mot de passe')
            
            flash(f'Les champs suivants sont obligatoires : {", ".join(missing_fields)}.', 'error')
            return redirect(url_for('register_student'))
            
        # Vérifier que les mots de passe correspondent
        if password != confirm_password:
            flash('Les mots de passe ne correspondent pas.', 'error')
            return redirect(url_for('register_student'))
        
        # Vérifier si l'email est déjà utilisé
        if User.query.filter_by(email=email).first():
            flash('Cet email est déjà utilisé.', 'error')
            return redirect(url_for('register_student'))
            
        # Vérifier si le nom d'utilisateur est déjà utilisé
        if User.query.filter_by(username=username).first():
            flash('Ce nom d\'utilisateur est déjà utilisé.', 'error')
            return redirect(url_for('register_student'))
        
        try:
            user = User(
                username=username,
                name=username,  # Utiliser le username comme nom par défaut
                email=email,
                role='student'
            )
            user.set_password(password)
            
            db.session.add(user)
            db.session.commit()
            app.logger.info("Nouvel utilisateur créé avec succès - username: %s, email: %s", username, email)
            flash('Inscription réussie ! Vous pouvez maintenant vous connecter.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            app.logger.error("Erreur lors de la création de l'utilisateur : %s", str(e))
            flash('Erreur lors de l\'inscription.', 'error')
            return redirect(url_for('register_student'))
    
    return render_template('register_student.html')

@app.route('/setup-admin')
def setup_admin():
    # Vérifier si un utilisateur existe déjà
    if User.query.first() is not None:
        flash('Un utilisateur existe déjà', 'warning')
        return redirect(url_for('login'))
    
    # Créer un compte administrateur par défaut
    admin = User(
        username='admin',
        name='Admin',
        email='admin@example.com',
        role='admin'
    )
    admin.set_password('admin')
    
    try:
        db.session.add(admin)
        db.session.commit()
        flash('Compte administrateur créé avec succès. Email: admin@example.com, Mot de passe: admin', 'success')
    except Exception as e:
        flash('Erreur lors de la création du compte administrateur', 'error')
    
    return redirect(url_for('login'))

@app.route('/dashboard')
@app.route('/teacher_dashboard')
@login_required
def teacher_dashboard():
    if not current_user.is_teacher:
        flash('Accès non autorisé.', 'error')
        return redirect(url_for('index'))
    
    classes = Class.query.filter_by(teacher_id=current_user.id).all()
    return render_template('teacher_dashboard.html', classes=classes)

@app.route('/statistics')
@login_required
def view_statistics():
    if not current_user.is_teacher:
        flash('Accès non autorisé.', 'error')
        return redirect(url_for('index'))
    
    classes = Class.query.filter_by(teacher_id=current_user.id).all()
    classes_stats = []
    
    for class_obj in classes:
        # Récupérer tous les élèves de la classe
        students = User.query.filter_by(class_id=class_obj.id, role='student').all()
        student_stats = []
        
        # Récupérer tous les exercices disponibles pour cette classe
        class_exercises = db.session.query(Exercise).join(ClassExercise).filter(ClassExercise.class_id == class_obj.id).all()
        total_exercises = len(class_exercises)
        
        # Pour chaque élève, calculer les statistiques
        for student in students:
            # Récupérer les tentatives d'exercices de l'élève
            attempts = ExerciseAttempt.query.filter_by(user_id=student.id).all()
            completed_exercises = len(set([a.exercise_id for a in attempts]))
            
            # Calculer le score moyen
            if attempts:
                average_score = sum([a.score for a in attempts]) / len(attempts)
            else:
                average_score = None
            
            student_stats.append({
                'student': student,
                'completed_exercises': completed_exercises,
                'total_exercises': total_exercises,
                'average_score': average_score
            })
        
        classes_stats.append({
            'class': class_obj,
            'students': student_stats,
            'total_exercises': total_exercises
        })
    
    return render_template('teacher/statistics.html', classes_stats=classes_stats)

@app.route('/teacher/export/pdf/<int:class_id>')
@login_required
def export_class_pdf(class_id):
    if not current_user.is_teacher:
        flash('Accès non autorisé.', 'error')
        return redirect(url_for('index'))
    
    # Vérifier que la classe appartient à l'enseignant
    class_obj = Class.query.filter_by(id=class_id, teacher_id=current_user.id).first_or_404()
    
    if not REPORTLAB_AVAILABLE:
        flash('L\'export PDF n\'est pas disponible. Veuillez installer ReportLab.', 'error')
        return redirect(url_for('view_statistics'))
    
    # Récupérer les données de la classe
    students = User.query.filter_by(class_id=class_obj.id, role='student').all()
    class_exercises = db.session.query(Exercise).join(ClassExercise).filter(ClassExercise.class_id == class_obj.id).all()
    total_exercises = len(class_exercises)
    
    # Préparer les données pour le PDF
    student_data = []
    for student in students:
        attempts = ExerciseAttempt.query.filter_by(user_id=student.id).all()
        completed_exercises = len(set([a.exercise_id for a in attempts]))
        
        if attempts:
            average_score = sum([a.score for a in attempts]) / len(attempts)
        else:
            average_score = 0
        
        student_data.append([
            student.name or student.username,
            completed_exercises,
            f"{average_score:.1f}%" if average_score else "0%",
            f"{(completed_exercises / total_exercises * 100):.0f}%" if total_exercises > 0 else "0%"
        ])
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Titre
    elements.append(Paragraph(f"Statistiques de la classe: {class_obj.name}", title_style))
    elements.append(Spacer(1, 12))
    
    # Informations générales
    elements.append(Paragraph(f"Enseignant: {current_user.name or current_user.username}", normal_style))
    elements.append(Paragraph(f"Date: {datetime.now().strftime('%d/%m/%Y')}", normal_style))
    elements.append(Paragraph(f"Nombre d'élèves: {len(students)}", normal_style))
    elements.append(Paragraph(f"Nombre d'exercices: {total_exercises}", normal_style))
    elements.append(Spacer(1, 12))
    
    # Tableau des élèves
    if student_data:
        elements.append(Paragraph("Résultats par élève", subtitle_style))
        elements.append(Spacer(1, 6))
        
        # En-têtes du tableau
        table_data = [["Nom de l'élève", "Exercices complétés", "Score moyen", "Progression"]]
        table_data.extend(student_data)
        
        # Créer le tableau
        table = Table(table_data)
        
        # Style du tableau
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        
        table.setStyle(table_style)
        elements.append(table)
    else:
        elements.append(Paragraph("Aucun élève dans cette classe.", normal_style))
    
    # Générer le PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Envoyer le PDF
    return send_from_directory(
        directory=os.path.dirname(os.path.abspath(__file__)),
        path=buffer,
        download_name=f"statistiques_{class_obj.name}_{datetime.now().strftime('%Y%m%d')}.pdf",
        as_attachment=True,
        mimetype='application/pdf'
    )

@app.route('/teacher/export/excel/<int:class_id>')
@login_required
def export_class_excel(class_id):
    if not current_user.is_teacher:
        flash('Accès non autorisé.', 'error')
        return redirect(url_for('index'))
    
    # Vérifier que la classe appartient à l'enseignant
    class_obj = Class.query.filter_by(id=class_id, teacher_id=current_user.id).first_or_404()
    
    if not OPENPYXL_AVAILABLE:
        flash('L\'export Excel n\'est pas disponible. Veuillez installer openpyxl.', 'error')
        return redirect(url_for('view_statistics'))
    
    # Récupérer les données de la classe
    students = User.query.filter_by(class_id=class_obj.id, role='student').all()
    class_exercises = db.session.query(Exercise).join(ClassExercise).filter(ClassExercise.class_id == class_obj.id).all()
    total_exercises = len(class_exercises)
    
    # Créer un classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Classe {class_obj.name}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    headers = ["Nom de l'élève", "Exercices complétés", "Score moyen", "Progression"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données des élèves
    row_num = 2
    for student in students:
        attempts = ExerciseAttempt.query.filter_by(user_id=student.id).all()
        completed_exercises = len(set([a.exercise_id for a in attempts]))
        
        if attempts:
            average_score = sum([a.score for a in attempts]) / len(attempts)
        else:
            average_score = 0
        
        progress = (completed_exercises / total_exercises * 100) if total_exercises > 0 else 0
        
        # Ajouter les données à la feuille
        ws.cell(row=row_num, column=1, value=student.name or student.username)
        ws.cell(row=row_num, column=2, value=completed_exercises)
        ws.cell(row=row_num, column=3, value=f"{average_score:.1f}%")
        ws.cell(row=row_num, column=4, value=f"{progress:.0f}%")
        
        row_num += 1
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Enregistrer le fichier Excel
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Envoyer le fichier Excel
    return send_from_directory(
        directory=os.path.dirname(os.path.abspath(__file__)),
        path=buffer,
        download_name=f"statistiques_{class_obj.name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/add_exercise_to_class/<int:exercise_id>', methods=['GET', 'POST'])
@login_required
def add_exercise_to_class(exercise_id):
    if current_user.role == 'student':
        return redirect(url_for('student_dashboard'))

    exercise = Exercise.query.get_or_404(exercise_id)
    # Note: Permettre aux enseignants d'ajouter des exercices d'autres enseignants à leurs classes
    # depuis la bibliothèque partagée (pas de restriction sur teacher_id pour l'ajout)

    # Récupérer les classes de l'enseignant
    classes = Class.query.filter_by(teacher_id=current_user.id).all()

    if request.method == 'POST':
        class_id = request.form.get('class_id')
        course_id = request.form.get('course_id')

        if not class_id or not course_id:
            flash('Veuillez sélectionner une classe et un cours', 'error')
            return redirect(url_for('add_exercise_to_class', exercise_id=exercise_id))

        # Vérifier que la classe et le cours existent et appartiennent à l'enseignant
        class_obj = Class.query.get_or_404(class_id)
        course = Course.query.get_or_404(course_id)

        if class_obj.teacher_id != current_user.id or course.class_obj.id != int(class_id):
            abort(403)

        # Ajouter l'exercice au cours s'il n'y est pas déjà
        if exercise not in course.exercises:
            course.exercises.append(exercise)
            db.session.commit()
            flash('L\'exercice a été ajouté au cours avec succès !', 'success')
        else:
            flash('L\'exercice est déjà dans ce cours', 'warning')

        return redirect(url_for('exercise_library'))

    return render_template('add_exercise_to_class.html', exercise=exercise, classes=classes)

@app.route('/get_courses/<int:class_id>')
@login_required
def get_courses(class_id):
    if current_user.role == 'student':
        return jsonify([]), 403

    # Vérifier que la classe appartient à l'enseignant
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.teacher_id != current_user.id:
        return jsonify([]), 403

    # Récupérer tous les cours de la classe
    courses = Course.query.filter_by(class_id=class_id).all()
    return jsonify([{'id': course.id, 'title': course.title} for course in courses])

@app.route('/class/<int:class_id>/delete', methods=['POST'])
@login_required
@teacher_required
def delete_class(class_id):
    logger.debug('='*50)
    logger.debug('TENTATIVE DE SUPPRESSION DE CLASSE')
    logger.debug(f'class_id: {class_id}')
    logger.debug(f'user: {current_user.name} (id: {current_user.id})')
    logger.debug(f'headers: {dict(request.headers)}')
    logger.debug(f'form data: {dict(request.form)}')
    
    # Vérification du token CSRF
    if not request.form.get('csrf_token'):
        logger.error("Token CSRF manquant")
        flash('Erreur de sécurité : token CSRF manquant.', 'error')
        return redirect(url_for('teacher_dashboard'))
    
    try:
        # Récupération et vérification de la classe
        class_obj = Class.query.get(class_id)
        if not class_obj:
            logger.error(f"Classe {class_id} non trouvée")
            flash('Classe non trouvée.', 'error')
            return redirect(url_for('teacher_dashboard'))
        
        # Vérification des permissions
        if class_obj.teacher_id != current_user.id:
            logger.error(f"L'utilisateur {current_user.name} n'est pas le professeur de la classe {class_obj.name}")
            flash('Vous n\'êtes pas autorisé à supprimer cette classe.', 'error')
            return redirect(url_for('teacher_dashboard'))
        
        try:
            # Supprimer les cours associés
            for course in class_obj.courses:
                db.session.delete(course)
            
            # Supprimer les associations avec les étudiants
            class_obj.students = []
            
            # Sauvegarder le nom pour le message
            class_name = class_obj.name
            
            # Supprimer la classe
            db.session.delete(class_obj)
            db.session.commit()
            
            logger.info(f"Classe {class_name} supprimée avec succès")
            flash(f'La classe {class_name} a été supprimée avec succès !', 'success')
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erreur SQL lors de la suppression: {str(e)}")
            flash('Une erreur est survenue lors de la suppression de la classe.', 'error')
            return redirect(url_for('view_class', class_id=class_id))
            
    except Exception as e:
        logger.error(f"Erreur inattendue: {str(e)}")
        flash('Une erreur inattendue est survenue.', 'error')
        
    return redirect(url_for('teacher_dashboard'))

@app.route('/course/<int:course_id>/delete', methods=['POST'])
@login_required
def delete_course(course_id):
    if not current_user.is_teacher:
        flash('Seuls les enseignants peuvent supprimer des cours.', 'error')
        return redirect(url_for('index'))
    
    course = Course.query.get_or_404(course_id)
    class_obj = Class.query.get_or_404(course.class_id)
    
    # Vérifier que l'enseignant est bien le propriétaire de la classe
    if class_obj.teacher_id != current_user.id:
        flash('Vous n\'avez pas la permission de supprimer ce cours.', 'error')
        return redirect(url_for('view_class', class_id=class_obj.id))
    
    try:
        # Supprimer les fichiers physiques
        for course_file in course.course_files:
            try:
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], course_file.filename)
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                app.logger.error(f"Erreur lors de la suppression du fichier {course_file.filename}: {str(e)}")
        
        # Supprimer le cours (les fichiers et exercices seront supprimés automatiquement grâce à cascade)
        db.session.delete(course)
        db.session.commit()
        flash('Le cours a été supprimé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Une erreur est survenue lors de la suppression du cours.', 'error')
        app.logger.error(f"Erreur lors de la suppression du cours {course_id}: {str(e)}")
    
    return redirect(url_for('view_class', class_id=class_obj.id))

@app.route('/class/<int:class_id>/remove-student/<int:student_id>', methods=['POST'])
@login_required
def remove_student_from_class(class_id, student_id):
    logger.debug('='*50)
    logger.debug('TENTATIVE DE SUPPRESSION D\'ÉTUDIANT')
    logger.debug(f'class_id: {class_id}')
    logger.debug(f'student_id: {student_id}')
    logger.debug(f'user: {current_user.name} (id: {current_user.id}, is_teacher: {current_user.is_teacher})')
    logger.debug(f'headers: {dict(request.headers)}')
    logger.debug(f'form data: {dict(request.form)}')
    
    # Vérification du token CSRF
    if not request.form.get('csrf_token'):
        logger.error("Token CSRF manquant")
        flash('Erreur de sécurité : token CSRF manquant.', 'error')
        return redirect(url_for('view_class', class_id=class_id))
    
    # Vérification des permissions
    if not current_user.is_teacher:
        logger.error(f"Accès refusé : l'utilisateur {current_user.name} n'est pas un enseignant")
        flash('Accès réservé aux enseignants.', 'error')
        return redirect(url_for('index'))
    
    try:
        # Vérification de la classe
        class_obj = Class.query.get(class_id)
        if not class_obj:
            logger.error(f"Classe {class_id} non trouvée")
            flash('Classe non trouvée.', 'error')
            return redirect(url_for('teacher_dashboard'))
            
        if class_obj.teacher_id != current_user.id:
            logger.error(f"L'utilisateur {current_user.name} n'est pas le professeur de la classe {class_obj.name}")
            flash('Vous n\'êtes pas le professeur de cette classe.', 'error')
            return redirect(url_for('teacher_dashboard'))
        
        # Vérification de l'étudiant
        student = User.query.get(student_id)
        if not student:
            logger.error(f"Étudiant {student_id} non trouvé")
            flash('Étudiant non trouvé.', 'error')
            return redirect(url_for('view_class', class_id=class_id))
            
        if not student in class_obj.students:
            logger.error(f"L'étudiant {student.name} n'est pas dans la classe {class_obj.name}")
            flash('Cet étudiant n\'est pas inscrit dans cette classe.', 'error')
            return redirect(url_for('view_class', class_id=class_id))
        
        try:
            # Supprimer l'étudiant de la classe
            class_obj.students.remove(student)
            db.session.commit()
            logger.info(f"Étudiant {student.name} supprimé avec succès de la classe {class_obj.name}")
            flash(f'L\'étudiant {student.name} a été retiré de la classe avec succès.', 'success')
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erreur SQL lors de la suppression: {str(e)}")
            flash('Une erreur est survenue lors de la suppression de l\'étudiant.', 'error')
            
    except Exception as e:
        logger.error(f"Erreur inattendue: {str(e)}")
        flash('Une erreur inattendue est survenue.', 'error')
        
    return redirect(url_for('view_class', class_id=class_id))

@app.route('/exercise-library')
@login_required
@teacher_required
def exercise_library():
    # Récupérer les paramètres de filtrage
    search_query = request.args.get('search', '')
    exercise_type = request.args.get('type', '')
    subject = request.args.get('subject', '')
    level = request.args.get('level', '')

    # Construire la requête de base
    query = Exercise.query

    # Appliquer les filtres
    if search_query:
        search = f"%{search_query}%"
        query = query.filter(
            db.or_(
                Exercise.title.ilike(search),
                Exercise.description.ilike(search)
            )
        )
    
    if exercise_type:
        query = query.filter(Exercise.exercise_type == exercise_type)
    
    # Exécuter la requête
    exercises = query.all()

    # Debug: afficher le nombre d'exercices trouvés
    app.logger.info(f"Nombre d'exercices trouvés : {len(exercises)}")
    for ex in exercises:
        app.logger.info(f"Exercice : {ex.title} (type: {ex.exercise_type})")

    # Debug: afficher les types d'exercices passés au template
    current_app.logger.debug(f"Types d'exercices passés au template: {Exercise.EXERCISE_TYPES}")
    
    # Debug supplémentaire: vérifier si "Mots à placer" est présent
    word_placement_present = any(type_id == 'word_placement' for type_id, type_name in Exercise.EXERCISE_TYPES)
    current_app.logger.debug(f"'Mots à placer' présent dans la liste: {word_placement_present}")
    
    # Debug: afficher chaque type individuellement
    for i, (type_id, type_name) in enumerate(Exercise.EXERCISE_TYPES):
        current_app.logger.debug(f"  {i+1}. {type_id} -> {type_name}")
    
    return render_template('teacher/exercise_library.html', 
                         exercises=exercises,
                         exercise_types=Exercise.EXERCISE_TYPES,
                         selected_search=search_query,
                         selected_type=exercise_type,
                         selected_subject=subject,
                         selected_level=level)

@app.route('/exercise/<int:exercise_id>')
@app.route('/exercise/<int:exercise_id>/<int:course_id>')
# @login_required  # TEMPORAIREMENT DÉSACTIVÉ POUR TEST
def view_exercise(exercise_id, course_id=None):
    try:
        app.logger.info(f'[VIEW_EXERCISE_DEBUG] Starting view_exercise for ID {exercise_id}')
        app.logger.info(f'Accessing exercise {exercise_id}')
        exercise = Exercise.query.get_or_404(exercise_id)
        app.logger.info(f'[VIEW_EXERCISE_DEBUG] Exercise type: {exercise.exercise_type}')
        app.logger.info(f'Found exercise: {exercise.title}')
        course_id = request.args.get('course_id', type=int)
        course = Course.query.get(course_id) if course_id else None
        app.logger.info(f'Course ID: {course_id}, Course: {course.title if course else None}')
    except Exception as e:
        app.logger.error(f'Error accessing exercise: {str(e)}')
        app.logger.exception('Full error:')
        return 'Une erreur est survenue lors de l\'accès à l\'exercice.', 500
    
    # Si l'exercice est accédé via un cours, vérifier que l'étudiant est inscrit
    # TEMPORAIREMENT DÉSACTIVÉ POUR TEST SANS AUTHENTIFICATION
    # if course and current_user.role == 'student':
    #     class_obj = Class.query.get(course.class_id)
    #     if not class_obj or current_user not in class_obj.students:
    #         flash('Vous n\'avez pas accès à cet exercice.', 'error')
    #         return redirect(url_for('index'))
    
    attempt = None
    user_answers = {}
    # Si l'utilisateur est un étudiant
    # TEMPORAIREMENT DÉSACTIVÉ POUR TEST SANS AUTHENTIFICATION
    # if current_user.role == 'student':
    #     # Récupérer la dernière tentative
    #     attempt = ExerciseAttempt.query.filter_by(
    #         exercise_id=exercise_id,
    #         student_id=current_user.id
    #     ).order_by(ExerciseAttempt.created_at.desc()).first()
    
    # Pour les tests sans authentification, récupérer la dernière tentative
    attempt = ExerciseAttempt.query.filter_by(
        exercise_id=exercise_id
    ).order_by(ExerciseAttempt.created_at.desc()).first()
    
    # Si une tentative existe, récupérer les réponses de l'utilisateur
    # mais seulement si l'exercice a été soumis (pour éviter d'afficher les réponses par défaut)
    show_answers = False
    if attempt:
        print(f'[VIEW_EXERCISE_DEBUG] Found attempt ID: {attempt.id}')
        # Vérifier si l'exercice a été soumis (présence de feedback)
        if attempt.feedback and attempt.feedback.strip():
            show_answers = True
            try:
                # Récupérer les réponses de l'utilisateur depuis le feedback
                feedback = json.loads(attempt.feedback) if attempt.feedback else []
                for item in feedback:
                    if 'student_answer' in item and 'blank_index' in item:
                        user_answers[f'answer_{item["blank_index"]}'] = item['student_answer']
                    elif 'student_answer' in item:
                        # Pour les autres formats de feedback
                        blank_counter = len(user_answers)
                        user_answers[f'answer_{blank_counter}'] = item['student_answer']
                print(f'[VIEW_EXERCISE_DEBUG] User answers: {user_answers}')
            except Exception as e:
                app.logger.error(f'Error parsing attempt feedback: {str(e)}')
                print(f'[VIEW_EXERCISE_DEBUG] Error parsing feedback: {str(e)}')
        else:
            print(f'[VIEW_EXERCISE_DEBUG] Attempt exists but no feedback, not showing answers')
            # Réinitialiser user_answers pour éviter d'afficher des réponses par défaut
            user_answers = {}
    
    # Les enseignants peuvent accéder directement aux exercices
    # TEMPORAIREMENT DÉSACTIVÉ POUR TEST SANS AUTHENTIFICATION
    # elif current_user.role != 'teacher':
    #     flash("Vous n'avez pas les permissions nécessaires.", "error")
    #     return redirect(url_for('index'))
        
    # Récupérer les statistiques et le contenu de l'exercice
    app.logger.info(f'Exercise type: {exercise.exercise_type}')
    app.logger.info(f'Raw content: {exercise.content}')
    try:
        content = exercise.get_content()
        app.logger.info(f'Parsed content: {content}')
        # Mélange des éléments de droite pour les appariements
        if exercise.exercise_type == 'pairs':
            # Support pour la nouvelle structure avec 'pairs'
            if 'pairs' in content:
                app.logger.info('Processing pairs with new structure')
                # Convertir la nouvelle structure vers l'ancienne pour compatibilité
                pairs = content.get('pairs', [])
                left_items = [pair['left'] for pair in pairs]
                right_items = [pair['right'] for pair in pairs]
                content['left_items'] = left_items
                content['right_items'] = right_items
                app.logger.info(f'Left items: {left_items}')
                app.logger.info(f'Right items: {right_items}')
                
                # Mélanger les éléments de droite
                import random
                shuffled = list(enumerate(right_items))
                random.shuffle(shuffled)
                content['shuffled_right_items'] = [item for idx, item in shuffled]
                content['shuffled_indices'] = [idx for idx, item in shuffled]
                app.logger.info(f'Shuffled right items: {content["shuffled_right_items"]}')
                app.logger.info(f'Shuffled indices: {content["shuffled_indices"]}')
            # Support pour l'ancienne structure avec 'right_items'
            elif 'right_items' in content:
                app.logger.info('Processing pairs with old structure')
                right_items = content.get('right_items', [])
                import random
                shuffled = list(enumerate(right_items))
                random.shuffle(shuffled)
                content['shuffled_right_items'] = [item for idx, item in shuffled]
                content['shuffled_indices'] = [idx for idx, item in shuffled]
            else:
                app.logger.error('No pairs or right_items found in content')
                app.logger.error(f'Content keys: {list(content.keys())}')
        
        # Traitement spécifique pour les exercices "Souligner les mots"
        elif exercise.exercise_type == 'underline_words':
            app.logger.info('Processing underline_words exercise')
            app.logger.info(f'Content keys: {list(content.keys())}')
            
            # Vérifier que les données nécessaires sont présentes
            if 'words' in content:
                words_data = content['words']
                app.logger.info(f'Found {len(words_data)} sentences in words data')
                
                for i, sentence_data in enumerate(words_data):
                    app.logger.info(f'Sentence {i}: {sentence_data.get("text", "NO TEXT")}')
                    app.logger.info(f'Words to underline {i}: {sentence_data.get("words_to_underline", "NO WORDS")}')
                
                # Ajouter les instructions si elles ne sont pas présentes
                if 'instructions' not in content:
                    content['instructions'] = 'Cliquez sur les mots à souligner dans les phrases ci-dessous.'
                    app.logger.info('Added default instructions')
            else:
                app.logger.error('No words data found in underline_words exercise')
                app.logger.error(f'Available keys: {list(content.keys())}')
                # Créer une structure vide pour éviter les erreurs
                content['words'] = []
                content['instructions'] = 'Aucune phrase trouvée pour cet exercice.'
        
        # Traitement spécifique pour les exercices "Mots mêlés"
        elif exercise.exercise_type == 'word_search':
            app.logger.info('Processing word_search exercise')
            app.logger.info(f'Content keys: {list(content.keys())}')
            
            # Vérifier que les données nécessaires sont présentes
            if 'words' in content and ('grid_size' in content or ('grid_width' in content and 'grid_height' in content)):
                words = content['words']
                
                # Support des deux formats de grille
                if 'grid_size' in content:
                    # Format ancien : {"grid_size": {"width": 15, "height": 15}}
                    grid_size = content['grid_size']
                    width = grid_size['width']
                    height = grid_size['height']
                else:
                    # Format nouveau : {"grid_width": 15, "grid_height": 15}
                    width = content['grid_width']
                    height = content['grid_height']
                    grid_size = {'width': width, 'height': height}
                
                app.logger.info(f'Found {len(words)} words to place in {width}x{height} grid')
                
                # Générer la grille avec les mots placés
                try:
                    from word_search_generator import generate_word_search_grid
                    grid, placed_words = generate_word_search_grid(words, grid_size['width'], grid_size['height'])
                    content['grid'] = grid
                    content['placed_words'] = placed_words
                    app.logger.info(f'Grid generated successfully with {len(placed_words)} words placed')
                except ImportError:
                    app.logger.warning('word_search_generator not available, using improved simple grid generation')
                    # Génération améliorée de grille
                    import random
                    import string
                    
                    width = grid_size['width']
                    height = grid_size['height']
                    
                    # Créer une grille vide
                    grid = []
                    for i in range(height):
                        row = []
                        for j in range(width):
                            row.append('')  # Grille vide au départ
                        grid.append(row)
                    
                    # Directions possibles : horizontal, vertical, diagonal
                    directions = [
                        (0, 1),   # horizontal droite
                        (1, 0),   # vertical bas
                        (1, 1),   # diagonal bas-droite
                        (-1, 1),  # diagonal haut-droite
                    ]
                    
                    placed_words = []
                    
                    # Fonction pour vérifier si un mot peut être placé
                    def can_place_word(word, start_row, start_col, direction):
                        dr, dc = direction
                        for i, letter in enumerate(word):
                            r = start_row + i * dr
                            c = start_col + i * dc
                            if r < 0 or r >= height or c < 0 or c >= width:
                                return False
                            if grid[r][c] != '' and grid[r][c] != letter:
                                return False
                        return True
                    
                    # Fonction pour placer un mot
                    def place_word(word, start_row, start_col, direction):
                        dr, dc = direction
                        for i, letter in enumerate(word):
                            r = start_row + i * dr
                            c = start_col + i * dc
                            grid[r][c] = letter
                        
                        end_row = start_row + (len(word) - 1) * dr
                        end_col = start_col + (len(word) - 1) * dc
                        
                        direction_name = {
                            (0, 1): 'horizontal',
                            (1, 0): 'vertical',
                            (1, 1): 'diagonal_down_right',
                            (-1, 1): 'diagonal_up_right'
                        }.get(direction, 'unknown')
                        
                        placed_words.append({
                            'word': word,
                            'start': [start_row, start_col],
                            'end': [end_row, end_col],
                            'direction': direction_name
                        })
                    
                    # Placer TOUS les mots (pas seulement 3)
                    for word in words:
                        placed = False
                        attempts = 0
                        max_attempts = 100
                        
                        while not placed and attempts < max_attempts:
                            # Choisir une direction aléatoire
                            direction = random.choice(directions)
                            dr, dc = direction
                            
                            # Calculer les limites de position de départ
                            if dr >= 0:
                                max_row = height - len(word)
                            else:
                                max_row = height - 1
                                min_row = len(word) - 1
                            
                            if dc >= 0:
                                max_col = width - len(word)
                            else:
                                max_col = width - 1
                                min_col = len(word) - 1
                            
                            # Générer une position aléatoire valide
                            if dr >= 0:
                                start_row = random.randint(0, max(0, max_row))
                            else:
                                start_row = random.randint(min_row, max_row)
                            
                            if dc >= 0:
                                start_col = random.randint(0, max(0, max_col))
                            else:
                                start_col = random.randint(min_col, max_col)
                            
                            # Vérifier si le mot peut être placé
                            if can_place_word(word, start_row, start_col, direction):
                                place_word(word, start_row, start_col, direction)
                                placed = True
                                app.logger.info(f'Placed word "{word}" at ({start_row}, {start_col}) direction {direction}')
                            
                            attempts += 1
                        
                        if not placed:
                            app.logger.warning(f'Could not place word "{word}" after {max_attempts} attempts')
                    
                    # Remplir les cases vides avec des lettres aléatoires
                    for i in range(height):
                        for j in range(width):
                            if grid[i][j] == '':
                                grid[i][j] = random.choice(string.ascii_uppercase)
                    
                    content['grid'] = grid
                    content['placed_words'] = placed_words
                    app.logger.info(f'Improved grid generated with {len(placed_words)} words placed out of {len(words)} total words')
                
                # Ajouter les instructions si elles ne sont pas présentes
                if 'instructions' not in content:
                    content['instructions'] = 'Trouvez tous les mots cachés dans la grille ci-dessous.'
                    app.logger.info('Added default instructions for word_search')
            else:
                app.logger.error('No words or grid_size found in word_search exercise')
                app.logger.error(f'Available keys: {list(content.keys())}')
                # Créer une structure vide pour éviter les erreurs
                content['words'] = []
                content['grid'] = []
                content['instructions'] = 'Aucune donnée trouvée pour cet exercice.'
        
        # Traitement spécifique pour les exercices "Mots à placer"
        elif exercise.exercise_type == 'word_placement':
            print(f'[WORD_PLACEMENT_DISPLAY] Processing word_placement exercise')
            print(f'[WORD_PLACEMENT_DISPLAY] Content keys: {list(content.keys())}')
            print(f'[WORD_PLACEMENT_DISPLAY] Raw content: {content}')
            
            # Vérifier que les données nécessaires sont présentes
            if 'sentences' in content and 'words' in content:
                sentences = content['sentences']
                words = content['words']
                print(f'[WORD_PLACEMENT_DISPLAY] Found {len(sentences)} sentences and {len(words)} words')
                
                # Ajouter les instructions si elles ne sont pas présentes
                if 'instructions' not in content:
                    content['instructions'] = 'Faites glisser les mots dans les bonnes phrases ou cliquez pour les placer.'
                    app.logger.info('Added default instructions for word_placement')
                
                # Pour word_placement, les phrases sont des chaînes simples avec des "___"
                # Pas besoin de traiter comme des dictionnaires avec blanks
                for i, sentence in enumerate(sentences):
                    if isinstance(sentence, str):
                        blank_count = sentence.count('___')
                        app.logger.info(f'Sentence {i}: "{sentence}" with {blank_count} blanks')
                    else:
                        app.logger.warning(f'Sentence {i} is not a string: {type(sentence)}')
                
                app.logger.info(f'Available words: {words}')
            else:
                app.logger.error('No sentences or words found in word_placement exercise')
                app.logger.error(f'Available keys: {list(content.keys())}')
                # Créer une structure vide pour éviter les erreurs
                content['sentences'] = []
                content['words'] = []
                content['instructions'] = 'Aucune donnée trouvée pour cet exercice.'
    except Exception as e:
        app.logger.error(f'Error parsing content: {str(e)}')
        app.logger.exception('Full error:')
        content = {'questions': []}
    progress = None
    
    # TEMPORAIREMENT DÉSACTIVÉ POUR TEST SANS AUTHENTIFICATION
    # progress = exercise.get_student_progress(current_user.id)
    
    # Choisir le template en fonction du type d'exercice
    if exercise.exercise_type == 'pairs':
        template = 'exercise_types/pairs_fixed.html'  # TEMPLATE CORRIGÉ
        app.logger.info(f'Using FIXED template for pairs: {template}')
    elif exercise.exercise_type == 'underline_words':
        template = 'exercise_types/underline_words.html'  # TEMPLATE ORIGINAL RESTAURÉ
        app.logger.info(f'Using ORIGINAL template for underline_words: {template}')
    else:
        template = f'exercise_types/{exercise.exercise_type}.html'
        app.logger.info(f'Using template: {template}')
    
    try:
        # Vérifier que le template existe
        if not os.path.exists(os.path.join(app.template_folder, template)):
            app.logger.error(f'Template not found: {template}')
            return 'Le template pour ce type d\'exercice n\'existe pas.', 500
        
        # Vérifier que les variables sont correctes
        app.logger.info(f'Variables for template:')
        app.logger.info(f'- exercise: {exercise}')
        app.logger.info(f'- attempt: {attempt}')
        app.logger.info(f'- content: {content}')
        app.logger.info(f'- progress: {progress}')
        app.logger.info(f'- course: {course}')
        
        return render_template(template,
                            exercise=exercise,
                            attempt=attempt,
                            content=content,
                            progress=progress,
                            course=course,
                            user_answers=user_answers,
                            show_answers=show_answers)
    except Exception as e:
        app.logger.error(f'Error rendering template: {str(e)}')
        app.logger.exception('Full template error:')
        return 'Une erreur est survenue lors de l\'affichage de l\'exercice.', 500

@app.route('/exercise/<int:exercise_id>/teacher')
@login_required
def view_exercise_teacher(exercise_id):
    if not current_user.is_teacher:
        flash("Accès non autorisé.", "error")
        return redirect(url_for('index'))
        
    exercise = Exercise.query.get_or_404(exercise_id)
    content = {}
    if exercise.content:
        content = json.loads(exercise.content)

    return render_template('view_exercise_teacher.html', exercise=exercise, content=content)

@app.route('/exercise/<int:exercise_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_exercise(exercise_id):
    if not current_user.is_teacher:
        flash("Accès non autorisé.", "error")
        return redirect(url_for('index'))
    
    exercise = Exercise.query.get_or_404(exercise_id)
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            title = request.form.get('title')
            description = request.form.get('description')
            subject = request.form.get('subject')
            exercise_type = request.form.get('exercise_type')
            
            # Mise à jour des champs de base
            exercise.title = title
            exercise.description = description
            exercise.subject = subject
            exercise.exercise_type = exercise_type
            
            # Gestion de l'image si elle est fournie
            if 'exercise_image' in request.files and request.files['exercise_image'].filename:
                image = request.files['exercise_image']
                if image and allowed_file(image.filename):
                    filename = secure_filename(image.filename)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    new_filename = f"{timestamp}_{filename}"
                    image_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
                    image.save(image_path)
                    exercise.image_path = new_filename
            
            # Mise à jour du contenu spécifique au type d'exercice
            content = {}
            if exercise.content:
                content = json.loads(exercise.content)
                
            # Traitement spécifique selon le type d'exercice
            if exercise_type == 'qcm':
                # Logique pour QCM
                questions = []
                for key, value in request.form.items():
                    if key.startswith('questions[') and key.endswith(']'):
                        question_index = int(key[10:-1])
                        while len(questions) <= question_index:
                            questions.append({"text": "", "choices": [], "correct": 0})
                        questions[question_index]["text"] = value
                    
                    elif key.startswith('options_'):
                        parts = key.split('_')
                        if len(parts) >= 2 and parts[1].isdigit():
                            question_index = int(parts[1])
                            while len(questions) <= question_index:
                                questions.append({"text": "", "choices": [], "correct": 0})
                            questions[question_index]["choices"].append(value)
                    
                    elif key.startswith('correct_'):
                        parts = key.split('_')
                        if len(parts) >= 2 and parts[1].isdigit():
                            question_index = int(parts[1])
                            while len(questions) <= question_index:
                                questions.append({"text": "", "choices": [], "correct": 0})
                            questions[question_index]["correct"] = int(value)
                
                content["questions"] = questions
                
            elif exercise_type == 'fill_in_blanks':
                # Logique pour texte à trous
                text = request.form.get('text', '')
                available_words = request.form.getlist('available_words[]')
                content = {"text": text, "available_words": available_words}
                
            # Autres types d'exercices...
            
            # Sauvegarde du contenu mis à jour
            exercise.content = json.dumps(content)
            
            # Sauvegarde des modifications
            db.session.commit()
            flash('Exercice mis à jour avec succès !', 'success')
            return redirect(url_for('view_exercise', exercise_id=exercise.id))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erreur lors de la mise à jour de l'exercice: {str(e)}")
            flash(f"Une erreur est survenue: {str(e)}", 'error')
    
    # Pour la méthode GET ou en cas d'erreur
    content = {}
    if exercise.content:
        content = json.loads(exercise.content)
    
    return render_template('edit_exercise.html', exercise=exercise, content=content)

@app.route('/course/<int:course_id>')
@login_required
def view_course(course_id):
    course = Course.query.get_or_404(course_id)

    # Vérifier que l'utilisateur a accès au cours
    if not current_user.is_teacher and not current_user.is_enrolled(course.class_obj.id):
        flash('Vous n\'avez pas accès à ce cours.', 'error')
        return redirect(url_for('index'))

    # Si c'est un enseignant, récupérer la liste des exercices disponibles
    exercises_available = []
    if current_user.is_teacher:
        # Récupérer tous les exercices créés par l'enseignant qui ne sont pas déjà dans le cours
        exercises_available = Exercise.query.filter_by(teacher_id=current_user.id).filter(
            ~Exercise.id.in_([ex.id for ex in course.exercises])
        ).all()

    # Récupérer les exercices du cours
    exercises = course.exercises

    # Pour les enseignants, récupérer les statistiques du cours
    stats = None
    if current_user.is_teacher:
        stats = {
            'total_students': len(course.class_obj.students),
            'total_exercises': len(exercises),
            'exercises_stats': []
        }


        for exercise in exercises:
            exercise_stats = {
                'title': exercise.title,
                'completion_rate': 0,
                'average_score': 0,
                'needs_grading': 0
            }

            total_students = len(course.class_obj.students)
            if total_students > 0:
                completed = 0
                total_score = 0
                needs_grading = 0

                for student in course.class_obj.students:
                    progress = exercise.get_student_progress(student.id)
                    if progress and progress.get('best_score') is not None:
                        completed += 1
                        total_score += progress['best_score']
                    elif progress and progress.get('needs_grading'):
                        needs_grading += 1

                exercise_stats['completion_rate'] = (completed / total_students) * 100
                if completed > 0:
                    exercise_stats['average_score'] = total_score / completed
                exercise_stats['needs_grading'] = needs_grading

            stats['exercises_stats'].append(exercise_stats)

    # Calcul de la progression pour les étudiants (partie droite)
    progress_records = []
    total_exercises = len(exercises)
    if not current_user.is_teacher:
        for ex in exercises:
            progress = ex.get_student_progress(current_user.id)
            if progress and progress.get('best_score', 0) >= 70:
                progress_records.append({'exercise_id': ex.id, 'student_id': current_user.id})

    return render_template('view_course.html',
                         course=course,
                         exercises=exercises,
                         exercises_available=exercises_available,
                         stats=stats,
                         progress_records=progress_records,
                         total_exercises=total_exercises)

@app.route('/class/<int:class_id>/create_course', methods=['GET', 'POST'])
@login_required
@teacher_required
def create_course(class_id):
    class_obj = Class.query.get_or_404(class_id)

    # Vérifier que l'utilisateur est bien le professeur de cette classe
    if class_obj.teacher_id != current_user.id:
        flash("Vous n'avez pas l'autorisation de créer un cours dans cette classe.", 'error')
        return redirect(url_for('teacher_dashboard'))

    class CourseForm(FlaskForm):
        title = StringField('Titre', validators=[DataRequired()])
        content = TextAreaField('Contenu')
        files = MultipleFileField('Fichiers joints')

    form = CourseForm()

    if form.validate_on_submit():
        # Récupérer le contenu de l'éditeur
        content = request.form.get('content', '{}')

        course = Course(
            title=form.title.data,
            content=json.dumps(content),  # Convertir en JSON
            class_id=class_id
        )

        # Gérer les fichiers
        files = request.files.getlist('files')
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                course_file = CourseFile(
                    filename=filename,
                    original_filename=file.filename,
                    file_type=file.content_type,
                    file_size=os.path.getsize(filepath),
                    course=course
                )
                db.session.add(course_file)

        db.session.add(course)
        db.session.commit()
        flash('Le cours a été créé avec succès.', 'success')
        return redirect(url_for('view_class', class_id=class_id))

    return render_template('create_course.html', form=form, class_id=class_id)

@app.route('/student/classes')
@login_required
def view_student_classes():
    if current_user.is_teacher:
        flash('Cette page est réservée aux étudiants.', 'error')
        return redirect(url_for('teacher_dashboard'))

    enrolled_classes = current_user.classes_enrolled
    return render_template('student_classes.html', classes=enrolled_classes)

@app.route('/class/join_by_code', methods=['GET', 'POST'])
@login_required
def join_class_by_code():
    if request.method == 'POST':
        class_code = request.form.get('class_code')
        if not class_code:
            flash('Le code d\'accès est requis.', 'error')
            return redirect(url_for('view_student_classes'))

        class_obj = Class.query.filter_by(access_code=class_code).first()
        if not class_obj:
            flash('Code d\'accès invalide.', 'error')
            return redirect(url_for('view_student_classes'))

        if current_user in class_obj.students:
            flash('Vous êtes déjà inscrit dans cette classe.', 'warning')
            return redirect(url_for('view_class', class_id=class_obj.id))

        try:
            class_obj.students.append(current_user)
            db.session.commit()
            flash('Vous avez rejoint la classe avec succès !', 'success')
            return redirect(url_for('view_class', class_id=class_obj.id))
        except Exception as e:
            db.session.rollback()
            flash('Une erreur est survenue lors de l\'inscription à la classe.', 'error')
            return redirect(url_for('view_student_classes'))

    return redirect(url_for('view_student_classes'))

@app.route('/class/<int:class_id>/view')
@login_required
def view_class(class_id):
    class_obj = Class.query.get_or_404(class_id)

    # Si c'est l'enseignant de la classe
    if current_user.is_teacher and class_obj.teacher_id == current_user.id:
        return view_class_teacher(class_id)

    # Si c'est un étudiant inscrit dans la classe
    if not current_user.is_teacher and class_obj in current_user.classes_enrolled:
        return render_template('view_class.html', class_obj=class_obj)

    flash('Accès non autorisé.', 'error')
    return redirect(url_for('index'))

@app.route('/class/<int:class_id>/view/teacher')
@login_required
@teacher_required
def view_class_teacher(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.teacher_id != current_user.id:
        flash('Vous n\'êtes pas le professeur de cette classe.', 'error')
        return redirect(url_for('index'))
    return render_template('view_class.html', class_obj=class_obj)

@app.route('/course/<int:course_id>/add-exercise', methods=['POST'])
@login_required
def add_exercise_to_course(course_id):
    app.logger.info(f"[add_exercise_to_course] Début de l'ajout d'exercice au cours {course_id}")
    app.logger.info(f"[add_exercise_to_course] Utilisateur: {current_user.id} ({current_user.role})")

    if not current_user.is_teacher:
        app.logger.warning("[add_exercise_to_course] Tentative d'accès non autorisé")
        flash('Accès non autorisé. Seuls les enseignants peuvent ajouter des exercices.', 'error')
        return redirect(url_for('index'))

    course = Course.query.get_or_404(course_id)
    app.logger.info(f"[add_exercise_to_course] Cours trouvé: {course.title}")

    # Vérifier que l'utilisateur est le propriétaire de la classe
    if course.class_obj.teacher_id != current_user.id:
        app.logger.warning("[add_exercise_to_course] L'utilisateur n'est pas le propriétaire de la classe")
        flash('Vous ne pouvez pas modifier ce cours.', 'error')
        return redirect(url_for('index'))

    exercise_id = request.form.get('exercise_id')
    app.logger.info(f"[add_exercise_to_course] ID de l'exercice reçu: {exercise_id}")

    if not exercise_id:
        app.logger.warning("[add_exercise_to_course] Aucun exercice sélectionné")
        flash('Veuillez sélectionner un exercice.', 'error')
        return redirect(url_for('view_course', course_id=course_id))

    exercise = Exercise.query.get_or_404(exercise_id)
    app.logger.info(f"[add_exercise_to_course] Exercice trouvé: {exercise.title}")

    # Vérifier que l'exercice n'est pas déjà dans le cours
    if exercise in course.exercises:
        app.logger.warning("[add_exercise_to_course] L'exercice est déjà dans le cours")
        flash('Cet exercice est déjà dans le cours.', 'error')
        return redirect(url_for('view_course', course_id=course_id))

    try:
        app.logger.info(f"[add_exercise_to_course] Tentative d'ajout de l'exercice {exercise_id} au cours {course_id}")
        app.logger.info(f"[add_exercise_to_course] État actuel du cours - Exercices: {[ex.id for ex in course.exercises]}")

        course.exercises.append(exercise)
        db.session.commit()

        app.logger.info(f"[add_exercise_to_course] Nouvel état du cours - Exercices: {[ex.id for ex in course.exercises]}")
        flash('Exercice ajouté au cours avec succès !', 'success')
        return redirect(url_for('view_course', course_id=course_id))
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[add_exercise_to_course] Erreur lors de l'ajout : {str(e)}")
        app.logger.error(f"[add_exercise_to_course] Type d'erreur : {type(e).__name__}")
        import traceback
        app.logger.error(f"[add_exercise_to_course] Traceback : {traceback.format_exc()}")
        flash('Erreur lors de l\'ajout de l\'exercice au cours.', 'error')
        return redirect(url_for('view_course', course_id=course_id))

@app.route('/course/<int:course_id>/remove-exercise/<int:exercise_id>', methods=['POST'])
@login_required
def remove_exercise_from_course(course_id, exercise_id):
    if not current_user.is_teacher:
        flash('Accès non autorisé. Seuls les enseignants peuvent retirer des exercices.', 'error')
        return redirect(url_for('index'))

    course = Course.query.get_or_404(course_id)

    # Vérifier que l'utilisateur est le propriétaire de la classe
    if course.class_obj.teacher_id != current_user.id:
        flash('Vous ne pouvez pas modifier ce cours.', 'error')
        return redirect(url_for('index'))

    exercise = Exercise.query.get_or_404(exercise_id)

    try:
        course.exercises.remove(exercise)
        db.session.commit()
        flash('Exercice retiré du cours avec succès !', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Erreur lors du retrait de l\'exercice du cours.', 'error')
        print(f"Erreur : {str(e)}")

    return redirect(url_for('view_course', course_id=course_id))

@app.route('/quick-add-exercise/<int:exercise_id>', methods=['GET', 'POST'])
@login_required
def process_quick_add_exercise(exercise_id):
    try:
        # Récupérer l'exercice
        exercise = Exercise.query.get_or_404(exercise_id)
        app.logger.info(f"[quick_add_exercise] Exercice trouvé: {exercise.title}")

        # Récupérer toutes les classes de l'utilisateur
        classes = Class.query.filter_by(teacher_id=current_user.id).all()
        app.logger.info(f"[quick_add_exercise] Nombre de classes trouvées: {len(classes)}")

        # Récupérer la classe sélectionnée si elle existe
        selected_class_id = request.args.get('class_id')
        selected_courses = []
        if selected_class_id:
            try:
                # Vérifier que la classe appartient à l'utilisateur
                class_obj = Class.query.get(int(selected_class_id))
                if class_obj and class_obj.teacher_id == current_user.id:
                    selected_courses = Course.query.filter_by(class_id=int(selected_class_id)).all()
                    app.logger.info(f"[quick_add_exercise] Cours trouvés pour la classe {selected_class_id}: {len(selected_courses)}")
            except ValueError:
                app.logger.error(f"[quick_add_exercise] ID de classe invalide: {selected_class_id}")
                selected_courses = []

        # Si c'est une requête POST, traiter l'ajout de l'exercice
        if request.method == 'POST':
            class_id = request.form.get('class_id')
            course_id = request.form.get('course_id')

            app.logger.info(f"[process_quick_add_exercise] Données reçues - class_id: {class_id}, course_id: {course_id}")

            if not class_id or not course_id:
                flash('Veuillez sélectionner une classe et un cours.', 'error')
                return redirect(url_for('process_quick_add_exercise', exercise_id=exercise_id, class_id=class_id))

            # Vérifier que la classe appartient à l'utilisateur
            class_obj = Class.query.get(class_id)
            if not class_obj or class_obj.teacher_id != current_user.id:
                flash('Classe non trouvée ou accès non autorisé.', 'error')
                return redirect(url_for('exercise_library'))

            # Vérifier que le cours appartient à la classe
            course = Course.query.get(course_id)
            if not course or course.class_id != int(class_id):
                flash('Cours non trouvé ou accès non autorisé.', 'error')
                return redirect(url_for('exercise_library'))

            # Ajouter l'exercice au cours s'il n'y est pas déjà
            if exercise not in course.exercises:
                app.logger.info(f"[process_quick_add_exercise] Ajout de l'exercice {exercise_id} au cours {course_id}")
                course.exercises.append(exercise)
                db.session.commit()
                flash('Exercice ajouté avec succès au cours !', 'success')
                return redirect(url_for('view_course', course_id=course_id))
            else:
                app.logger.info(f"[process_quick_add_exercise] L'exercice {exercise_id} est déjà dans le cours {course_id}")
                flash('Cet exercice est déjà dans le cours.', 'info')
                return redirect(url_for('view_course', course_id=course_id))

        # Pour les requêtes GET, afficher le formulaire
        return render_template('add_exercise_to_class.html',
                             exercise=exercise,
                             classes=classes,
                             selected_class_id=selected_class_id,
                             selected_courses=selected_courses)

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[process_quick_add_exercise] Erreur lors de l'ajout de l'exercice: {str(e)}")
        app.logger.error(f"[process_quick_add_exercise] Type d'erreur: {type(e).__name__}")
        import traceback
        app.logger.error(f"[process_quick_add_exercise] Traceback: {traceback.format_exc()}")
        flash('Une erreur est survenue lors de l\'ajout de l\'exercice.', 'error')
        return redirect(url_for('process_quick_add_exercise', exercise_id=exercise_id))

@app.route('/class/create', methods=['GET', 'POST'])
@login_required
@teacher_required
def create_class():
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')

        if not name:
            flash('Le nom de la classe est requis.', 'error')
            return redirect(url_for('create_class'))

        # Générer un code d'accès unique
        import random
        import string

        access_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        while Class.query.filter_by(access_code=access_code).first() is not None:
            access_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

        try:
            new_class = Class(
                name=name,
                description=description,
                teacher_id=current_user.id,
                access_code=access_code
            )

            db.session.add(new_class)
            db.session.commit()

            # Afficher le code d'accès à l'enseignant
            flash(f'Classe créée avec succès ! Code d\'accès : {access_code}', 'success')
            return redirect(url_for('view_class', class_id=new_class.id))
        except Exception as e:
            db.session.rollback()
            flash('Une erreur est survenue lors de la création de la classe.', 'error')
            return redirect(url_for('create_class'))

    return render_template('create_class.html')

@app.route('/class/<int:class_id>/edit', methods=['GET', 'POST'])
@login_required
@teacher_required
def edit_class(class_id):
    class_obj = Class.query.get_or_404(class_id)

    # Vérifier que l'utilisateur est bien le propriétaire de la classe
    if class_obj.teacher_id != current_user.id:
        flash('Vous n\'avez pas la permission de modifier cette classe.', 'error')
        return redirect(url_for('teacher_dashboard'))

    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')

        if not name:
            flash('Le nom de la classe est requis.', 'error')
            return redirect(url_for('edit_class', class_id=class_id))

        try:
            class_obj.name = name
            class_obj.description = description
            db.session.commit()
            flash('Classe modifiée avec succès !', 'success')
            return redirect(url_for('view_class', class_id=class_id))
        except Exception as e:
            db.session.rollback()
            flash('Une erreur est survenue lors de la modification de la classe.', 'error')
            return redirect(url_for('edit_class', class_id=class_id))

    return render_template('edit_class.html', class_obj=class_obj)

@app.route('/course/<int:course_id>/edit', methods=['GET', 'POST'])
@login_required
@teacher_required
def edit_course(course_id):
    course = Course.query.get_or_404(course_id)
    class_obj = Class.query.get_or_404(course.class_id)

    # Vérifier que l'utilisateur est le professeur de la classe
    if current_user.id != class_obj.teacher_id:
        flash("Vous n'êtes pas autorisé à modifier ce cours.", 'error')
        return redirect(url_for('view_class', class_id=class_obj.id))

    class CourseForm(FlaskForm):
        title = StringField('Titre', validators=[DataRequired()])
        files = MultipleFileField('Ajouter des fichiers')

    form = CourseForm(obj=course)

    if form.validate_on_submit():
        course.title = form.title.data
        course.content = json.dumps(request.form.get('content', '{}'))  # Convertir en JSON

        # Gérer les nouveaux fichiers
        files = request.files.getlist('files')
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                course_file = CourseFile(
                    filename=filename,
                    original_filename=file.filename,
                    file_type=file.content_type,
                    file_size=os.path.getsize(filepath),
                    course=course
                )
                db.session.add(course_file)

        db.session.commit()
        flash('Le cours a été modifié avec succès !', 'success')
        return redirect(url_for('view_course', course_id=course.id))

    return render_template('edit_course.html', course=course, form=form)

@app.route('/course/<int:course_id>/file/<int:file_id>/delete', methods=['POST'])
@login_required
@teacher_required
def delete_course_file(course_id, file_id):
    course = Course.query.get_or_404(course_id)
    course_file = CourseFile.query.get_or_404(file_id)

    # Vérifier que le fichier appartient bien au cours
    if course_file.course_id != course.id:
        return jsonify({'error': 'Ce fichier n\'appartient pas à ce cours.'}), 403

    # Vérifier que l'utilisateur est bien le professeur de la classe
    class_obj = Class.query.get(course.class_id)
    if current_user.id != class_obj.teacher_id:
        return jsonify({'error': 'Vous n\'êtes pas autorisé à supprimer ce fichier.'}), 403

    try:
        # Supprimer le fichier physique
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], course_file.filename)
        if os.path.exists(file_path):
            os.remove(file_path)

        # Supprimer l'entrée dans la base de données
        db.session.delete(course_file)
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/exercise/underline-words-example')
@app.route('/exercise/<int:exercise_id>/submit', methods=['POST'])
@login_required
def submit_exercise(exercise_id):
    exercise = Exercise.query.get_or_404(exercise_id)
    course_id = request.form.get('course_id')
    
    # Si un course_id est fourni, vérifier les permissions
    if course_id:
        course = Course.query.get_or_404(course_id)
        if not current_user.is_enrolled(course.class_obj.id):
            flash('Vous n\'avez pas accès à cet exercice.', 'error')
            return redirect(url_for('student_classes'))
        
        # Vérifier que l'exercice fait partie du cours
        if exercise not in course.exercises:
            flash('Cet exercice ne fait pas partie du cours.', 'error')
            return redirect(url_for('view_course', course_id=course_id))
    
    # Vérifier le nombre de tentatives restantes
    progress = exercise.get_student_progress(current_user.id)
    if exercise.max_attempts and progress and progress['remaining_attempts'] is not None and progress['remaining_attempts'] <= 0:
        flash('Vous avez atteint le nombre maximum de tentatives pour cet exercice.', 'error')
        return redirect(url_for('view_exercise', exercise_id=exercise_id))
        return redirect(url_for('view_exercise', exercise_id=exercise_id, course_id=course_id))
    
    # Traiter les réponses
    answers = []
    score = 0
    feedback = []
    content = exercise.get_content()
    
    if exercise.exercise_type == 'qcm':
        total_questions = len(content['questions'])
        correct_answers = 0
        
        app.logger.debug(f'Content: {content}')
        for i, question in enumerate(content['questions']):
            answer = request.form.get(f'answer_{i}')
            app.logger.debug(f'Réponse brute pour la question {i}: {answer}')
            
            # Convertir la réponse en entier
            try:
                answer = int(answer) if answer is not None else -1
            except (ValueError, TypeError):
                answer = -1
            answers.append(answer)
            
            # Accepter à la fois 'correct' et 'correct_answer' comme clé de la bonne réponse
            correct_key = 'correct'
            if 'correct' not in question and 'correct_answer' in question:
                correct_key = 'correct_answer'
            if correct_key not in question:
                app.logger.error(f"La question {i} ne contient pas la clé 'correct' ou 'correct_answer'. Structure de la question : {question}")
                feedback.append({
                    'question': i + 1,
                    'correct': False,
                    'message': "Erreur interne : la question est mal structurée. Merci de contacter l'enseignant."
                })
                continue  # On passe à la question suivante
            # S'assurer que la valeur est bien un entier pour la comparaison
            try:
                correct = int(question[correct_key])
            except (ValueError, TypeError, KeyError):
                correct = -1
            app.logger.debug(f'Question {i} - Réponse de l\'utilisateur: {answer}, Réponse correcte: {correct}')
            
            # Forcer la conversion en int pour éviter les erreurs de type
            try:
                answer_int = int(answer)
            except (ValueError, TypeError):
                app.logger.warning(f"Impossible de convertir la réponse utilisateur '{answer}' en int pour la question {i}.")
                answer_int = -999
            try:
                correct_int = int(correct)
            except (ValueError, TypeError):
                app.logger.warning(f"Impossible de convertir la réponse correcte '{correct}' en int pour la question {i}.")
                correct_int = -999
            # Vérifier si la réponse est correcte
            if answer_int == correct_int:
                correct_answers += 1
                app.logger.debug(f'Question {i} - Bonne réponse !')
                feedback.append({
                    'question': i + 1,
                    'correct': True,
                    'message': 'Bonne réponse !'
                })
            else:
                # Récupérer le texte de la réponse correcte
                correct_option = question['options'][correct] if 0 <= correct < len(question['options']) else 'Non spécifiée'
                app.logger.debug(f'Question {i} - Mauvaise réponse. Attendu: {correct_option}')
                feedback.append({
                    'question': i + 1,
                    'correct': False,
                    'message': f'La réponse correcte était : {correct_option}'
                })
        
        app.logger.debug(f'Réponses correctes: {correct_answers}, Total questions: {total_questions}')
        score = (correct_answers / total_questions) * 100 if total_questions > 0 else 0
        app.logger.debug(f'Score final: {score}%')
    
    elif exercise.exercise_type == 'fill_in_blanks':
        # Cette implémentation a été désactivée car elle est en conflit avec l'implémentation plus complète
        # Voir lignes ~3415-3510 pour l'implémentation active.
        pass

    elif exercise.exercise_type == 'word_placement':
        user_answers = []
        correct_answers = content.get('answers', [])
        
        # Récupérer toutes les réponses de l'utilisateur dans l'ordre séquentiel
        i = 0
        while True:
            answer = request.form.get(f'answer_{i}')
            if answer is None:
                break
            user_answers.append(answer.strip())
            i += 1
        
        app.logger.debug(f'[WORD_PLACEMENT_SCORING] Réponses utilisateur: {user_answers}')
        app.logger.debug(f'[WORD_PLACEMENT_SCORING] Réponses correctes: {correct_answers}')
        
        if len(user_answers) != len(correct_answers):
            flash('Nombre de réponses incorrect.', 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id))
        
        score = 0
        feedback = []
        for i, (user_ans, correct_ans) in enumerate(zip(user_answers, correct_answers)):
            # Comparaison insensible à la casse et aux espaces
            is_correct = user_ans.strip().lower() == correct_ans.strip().lower()
            score += 1 if is_correct else 0
            feedback.append({
                'user_answer': user_ans,
                'correct_answer': correct_ans,
                'is_correct': is_correct,
                'position': i + 1
            })
        
        max_score = len(correct_answers)
        score = (score / max_score) * 100 if max_score > 0 else 0
        
        app.logger.debug(f'[WORD_PLACEMENT_SCORING] Score final: {score}% ({score}/{max_score})')
    
    elif exercise.exercise_type == 'word_search':
        found_words = request.form.getlist('found_words[]')
        total_words = len(content['words'])
        correct_words = 0
        
        # Convertir les mots trouvés et les mots à chercher en minuscules pour la comparaison
        found_words = [word.lower().strip() for word in found_words]
        target_words = [word.lower().strip() for word in content['words']]
        
        for word in found_words:
            if word in target_words:
                correct_words += 1
                feedback.append({
                    'word': word,
                    'correct': True,
                    'message': 'Mot trouvé !'
                })
        
        # Ajouter le feedback pour les mots non trouvés
        for word in target_words:
            if word not in found_words:
                feedback.append({
                    'word': word,
                    'correct': False,
                    'message': 'Mot non trouvé'
                })
        
        score = (correct_words / total_words) * 100 if total_words > 0 else 0
        
    elif exercise.exercise_type == 'pairs':
        # Structure attendue : 'correct_pairs': [[0, 0], [1, 1], ...], left_items, right_items
        correct_pairs = content.get('correct_pairs', [])
        total_pairs = len(correct_pairs)
        correct_count = 0
        user_pairs = []
        for i in range(total_pairs):
            left = request.form.get(f'left_{i}')
            right = request.form.get(f'right_{i}')
            if left is not None and right is not None:
                try:
                    left_idx = int(left)
                    right_idx = int(right)
                    user_pairs.append([left_idx, right_idx])
                    answers.append({'left': left_idx, 'right': right_idx})
                except (ValueError, TypeError):
                    continue
        for pair in correct_pairs:
            if pair in user_pairs:
                correct_count += 1
                feedback.append({'pair': pair, 'correct': True, 'message': 'Paire correcte !'})
            else:
                feedback.append({'pair': pair, 'correct': False, 'message': 'Paire incorrecte'})
        score = (correct_count / total_pairs) * 100 if total_pairs > 0 else 0
    
    # S'assurer que le score est un nombre valide
    if score is None:
        score = 0.0
    
    # Enregistrer la tentative
    attempt = ExerciseAttempt(
        student_id=current_user.id,
        exercise_id=exercise_id,
        course_id=course_id,
        score=float(score),  # Convertir explicitement en float
        answers=json.dumps(answers),
        feedback=json.dumps(feedback),
        completed=True
    )
    
    db.session.add(attempt)
    db.session.commit()
    
    # Rafraîchir l'objet depuis la base de données pour s'assurer que tout est à jour
    db.session.refresh(attempt)
    
    flash(f'Exercice soumis avec succès ! Score : {score:.1f}%', 'success')
    return redirect(url_for('view_exercise', exercise_id=exercise_id, course_id=course_id))

@app.route('/exercise/<int:exercise_id>/stats')
@login_required
@teacher_required
def exercise_stats(exercise_id):
    exercise = Exercise.query.get_or_404(exercise_id)
    course_id = request.args.get('course_id', type=int)
    
    # Vérifier que l'enseignant a le droit d'accéder à ces statistiques
    has_access = False
    
    # Vérifier si l'enseignant est le créateur de l'exercice
    if exercise.teacher_id == current_user.id:
        has_access = True
    else:
        # Vérifier si l'exercice est utilisé dans l'une des classes de l'enseignant
        teacher_classes = Class.query.filter_by(teacher_id=current_user.id).all()
        for class_obj in teacher_classes:
            for course in class_obj.courses:
                if exercise in course.exercises:
                    has_access = True
                    break
            if has_access:
                break
    
    if not has_access:
        flash("Vous n'avez pas l'autorisation de voir ces statistiques.", "error")
        return redirect(url_for('index'))
    
    # Récupérer les statistiques
    stats = exercise.get_stats(course_id)
    
    # Récupérer les informations du cours si spécifié
    course = Course.query.get(course_id) if course_id else None
    
    # Ajouter les informations de progression pour chaque étudiant
    student_progress = []
    
    if course:
        # Pour un cours spécifique, ne montrer que les étudiants de la classe associée
        students_to_show = course.class_obj.students
    else:
        # Sans cours spécifié, montrer tous les étudiants qui ont déjà fait l'exercice
        attempts = ExerciseAttempt.query.filter_by(exercise_id=exercise.id).all()
        student_ids = list(set([a.student_id for a in attempts]))  # Utiliser set() pour éliminer les doublons
        students_to_show = User.query.filter(User.id.in_(student_ids), User.role=='student').all()
    
    for student in students_to_show:
        progress = exercise.get_student_progress(student.id)
        if progress:
            student_progress.append({
                'student': student,
                'progress': progress
            })
    
    return render_template('exercise_stats.html',
                         exercise=exercise,
                         course=course,
                         stats=stats,
                         student_progress=student_progress)

@app.route('/exercise/<int:exercise_id>/feedback/<int:attempt_id>')
@login_required
def view_feedback(exercise_id, attempt_id):
    exercise = Exercise.query.get_or_404(exercise_id)
    attempt = ExerciseAttempt.query.get_or_404(attempt_id)
    
    # Vérifier que l'utilisateur a le droit de voir ce feedback
    if not current_user.is_teacher and attempt.student_id != current_user.id:
        flash("Vous n'avez pas l'autorisation de voir cette tentative.", "error")
        return redirect(url_for('index'))

    # Si c'est un enseignant, vérifier qu'il enseigne dans la classe associée au cours
    if current_user.is_teacher and attempt.course_id:
        course = Course.query.get(attempt.course_id)
        if course and course.class_obj.teacher_id != current_user.id:
            flash("Vous n'avez pas l'autorisation de voir cette tentative.", "error")
            return redirect(url_for('index'))
    
    # Convertir les réponses et le feedback en dictionnaires
    answers = {}
    feedback = {}
    
    if exercise.exercise_type == 'qcm':
        answers_list = json.loads(attempt.answers)
        for i, answer in enumerate(answers_list, 1):
            answers[str(i)] = answer
    elif exercise.exercise_type == 'pair_match':
        pairs = json.loads(attempt.answers)
        for i, pair in enumerate(pairs, 1):
            answers[str(i)] = f"{pair['left']} → {pair['right']}"
    
    if attempt.feedback:
        feedback = json.loads(attempt.feedback)
    
    return render_template('feedback.html',
                         exercise=exercise,
                         attempt=attempt,
                         answers=answers,
                         feedback=feedback)

@app.route('/course/<int:course_id>/file/<int:file_id>/download')
@login_required
def download_course_file(course_id, file_id):
    course = Course.query.get_or_404(course_id)
    file = CourseFile.query.get_or_404(file_id)
    
    if file.course_id != course.id:
        flash('Fichier non trouvé.', 'error')
        return redirect(url_for('view_course', course_id=course_id))
    
    # Vérifier que l'utilisateur a accès au cours
    if current_user.is_teacher or any(c.id == course.class_id for c in current_user.classes_enrolled):
        uploads_dir = app.config['UPLOAD_FOLDER']  # Utilise le bon dossier static/uploads
        return send_from_directory(uploads_dir, file.filename, as_attachment=True, download_name=file.original_filename)
    else:
        flash('Accès non autorisé.', 'error')
        return redirect(url_for('index'))

@app.route('/course/<int:course_id>/get-available-exercises')
@login_required
def get_available_exercises(course_id):
    """Retourne la liste des exercices disponibles pour un cours"""
    if not current_user.is_teacher:
        return jsonify({'error': 'Non autorisé'}), 403
    
    course = Course.query.get_or_404(course_id)
    if course.class_obj.teacher_id != current_user.id:
        return jsonify({'error': 'Non autorisé'}), 403
    
    # Récupérer tous les exercices disponibles
    exercises = Exercise.query.all()
    
    # Filtrer les exercices qui ne sont pas déjà dans le cours
    available_exercises = [
        {'id': ex.id, 'title': ex.title, 'type': ex.exercise_type}
        for ex in exercises if ex not in course.exercises
    ]
    
    return jsonify(available_exercises)

@app.route('/exercise/<int:exercise_id>/submit/<int:course_id>', methods=['POST'])
# @login_required  # Commenté pour les tests sans authentification

def submit_answer(exercise_id, course_id=0):
    app.logger.info(f"[DEBUG] submit_exercise_answer called for exercise_id={exercise_id}")
    print("\n=== DÉBUT SUBMIT_ANSWER ===\n")
    print(f"[DEBUG] Soumission pour l'exercice {exercise_id}")
    # print(f"[DEBUG] Utilisateur: {current_user.username} (ID: {current_user.id})")  # Commenté pour tests sans auth
    print("[DEBUG] Form data:", request.form)
    print("[DEBUG] Form data type:", type(request.form))
    
    exercise = Exercise.query.get_or_404(exercise_id)
    # Utiliser le course_id de l'URL ou du formulaire comme fallback
    if not course_id:
        course_id = request.form.get('course_id')
    print(f"[DEBUG] Course ID from route: {course_id}")
    print(f"[DEBUG] Course ID from form: {request.form.get('course_id')}")
    
    # if not course_id:  # Commenté pour tests sans cours
    #     flash('Erreur: Cours non spécifié', 'error')
    #     return redirect(url_for('exercise_library'))
    
    # Vérifier que l'étudiant a accès à ce cours
    # course = Course.query.get_or_404(course_id)  # Commenté pour tests sans cours
    course = None  # Pour les tests
    # if not current_user.is_enrolled(course.class_obj.id):  # Commenté pour tests sans auth
    #     flash('Vous n\'avez pas accès à cet exercice.', 'error')
    #     return redirect(url_for('exercise_library'))
    
    answers = {}
    score = 0
    feedback = []

    if exercise.exercise_type == 'qcm':
        print("\n=== DÉBUT SCORING QCM ===")
        content = exercise.get_content()
        print(f"[QCM_DEBUG] Content: {content}")
        
        if not isinstance(content, dict) or 'questions' not in content:
            print("[QCM_DEBUG] Structure invalide!")
            flash('Structure de l\'exercice invalide.', 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id))

        total_questions = len(content['questions'])
        correct_answers = 0
        print(f"[QCM_DEBUG] Total questions: {total_questions}")

        for i, question in enumerate(content['questions']):
            student_answer = request.form.get(f'answer_{i}')
            correct_answer = question.get('correct_answer')
            
            print(f"[QCM_DEBUG] Question {i}:")
            print(f"  - Texte: {question.get('text')}")
            print(f"  - Réponse étudiant (answer_{i}): {student_answer}")
            print(f"  - Bonne réponse: {correct_answer}")
            print(f"  - Choices: {question.get('choices')}")
            
            is_correct = False
            if student_answer is not None and correct_answer is not None:
                is_correct = int(student_answer) == int(correct_answer)
                if is_correct:
                    correct_answers += 1
                    print(f"  - [CORRECT] Bonne reponse!")
                else:
                    print(f"  - [INCORRECT] {student_answer} != {correct_answer}")
            else:
                print(f"  - [WARNING] DONNEES MANQUANTES: student={student_answer}, correct={correct_answer}")

            feedback.append({
                'question': question['text'],
                'student_answer': question['choices'][int(student_answer)] if student_answer and int(student_answer) < len(question['choices']) else 'Aucune réponse',
                'correct_answer': question['choices'][correct_answer] if correct_answer < len(question['choices']) else 'Erreur',
                'is_correct': is_correct
            })

        score = (correct_answers / total_questions) * 100 if total_questions > 0 else 0
        print(f"[QCM_DEBUG] Score final: {correct_answers}/{total_questions} = {score}%")
        print("=== FIN SCORING QCM ===\n")

    elif exercise.exercise_type == 'underline_words':
        content = exercise.get_content()
        
        # Gérer les deux structures possibles : 'sentences' ou 'words'
        sentences_data = content.get('sentences') or content.get('words', [])
        
        if not isinstance(content, dict) or not sentences_data:
            flash('Structure de l\'exercice invalide.', 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id, course_id=course_id))

        total_sentences = len(sentences_data)
        correct_sentences = 0

        for i, sentence_data in enumerate(sentences_data):
            student_answers = request.form.getlist(f'underlined_words_{i}[]')
            correct_words = set(word.lower() for word in sentence_data['words_to_underline'])
            student_words = set(word.lower() for word in student_answers)

            is_correct = student_words == correct_words
            if is_correct:
                correct_sentences += 1

            feedback.append({
                'sentence': sentence_data['text'],
                'student_words': list(student_words),
                'correct_words': list(correct_words),
                'is_correct': is_correct
            })

        score = (correct_sentences / total_sentences) * 100 if total_sentences > 0 else 0

    elif exercise.exercise_type == 'pairs':
        content = exercise.get_content()
        if not isinstance(content, dict) or 'pairs' not in content:
            flash('Structure de l\'exercice invalide.', 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id, course_id=course_id))

        total_pairs = len(content['pairs'])
        correct_pairs = 0

        # Créer un dictionnaire des paires correctes
        correct_matches = {pair['first']: pair['second'] for pair in content['pairs']}

        for first, correct_second in correct_matches.items():
            student_answer = request.form.get(f'pair_{first}')
            is_correct = student_answer == correct_second

            if is_correct:
                correct_pairs += 1

            feedback.append({
                'first': first,
                'student_answer': student_answer,
                'correct_answer': correct_second,
                'is_correct': is_correct
            })

        score = (correct_pairs / total_pairs) * 100 if total_pairs > 0 else 0

    elif exercise.exercise_type == 'fill_in_blanks':
        content = exercise.get_content()
        print("[DEBUG] Contenu de l'exercice:", content)

        # Vérification de la structure attendue
        if not isinstance(content, dict) or 'sentences' not in content or not isinstance(content['sentences'], list):
            app.logger.error(f"[ERREUR STRUCTURE] L'exercice {exercise.id} n'a pas de clé 'sentences' valide dans son contenu: {content}")
            flash("Erreur de configuration de l'exercice : structure des phrases manquante ou invalide.", 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id, course_id=course_id))

        total_questions = len(content['sentences'])
        correct_answers = 0

        for i in range(total_questions):
            student_answer = request.form.get(f'answer_{i}')
            correct_answer = content['sentences'][i]['answer']
            print(f"[DEBUG] Question {i + 1}:")
            print(f"[DEBUG] - Réponse de l'étudiant : {student_answer}")
            print(f"[DEBUG] - Réponse correcte : {correct_answer}")

            if student_answer and student_answer.strip().lower() == correct_answer.strip().lower():
                correct_answers += 1
                feedback.append({
                    'question': content['sentences'][i]['text'],
                    'student_answer': student_answer,
                    'correct_answer': correct_answer,
                    'is_correct': True
                })
            else:
                feedback.append({
                    'question': content['sentences'][i]['text'],
                    'student_answer': student_answer or '',
                    'correct_answer': correct_answer,
                    'is_correct': False
                })

        score = (correct_answers / total_questions) * 100 if total_questions > 0 else 0

    elif exercise.exercise_type == 'word_placement':
        print("\n=== DÉBUT SCORING WORD_PLACEMENT ===")
        content = exercise.get_content()
        print(f"[WORD_PLACEMENT_DEBUG] Content: {content}")
        
        if not isinstance(content, dict) or 'sentences' not in content or 'answers' not in content:
            print("[WORD_PLACEMENT_DEBUG] Structure invalide!")
            flash('Structure de l\'exercice invalide.', 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id))

        sentences = content['sentences']
        correct_answers = content['answers']
        
        # CORRECTION: Compter le nombre réel de blancs dans les phrases
        total_blanks_in_sentences = sum(s.count('___') for s in sentences)
        total_blanks = max(total_blanks_in_sentences, len(correct_answers))
        
        correct_count = 0
        
        print(f"[WORD_PLACEMENT_DEBUG] Total blanks in sentences: {total_blanks_in_sentences}")
        print(f"[WORD_PLACEMENT_DEBUG] Total answers: {len(correct_answers)}")
        print(f"[WORD_PLACEMENT_DEBUG] Using total_blanks = {total_blanks}")
        print(f"[WORD_PLACEMENT_DEBUG] Expected answers: {correct_answers}")

        # Vérifier chaque réponse
        for i in range(total_blanks):
            student_answer = request.form.get(f'answer_{i}')
            expected_answer = correct_answers[i] if i < len(correct_answers) else ''
            
            print(f"[WORD_PLACEMENT_DEBUG] Blank {i}:")
            print(f"  - Réponse étudiant (answer_{i}): {student_answer}")
            print(f"  - Réponse attendue: {expected_answer}")
            
            if student_answer and student_answer.strip().lower() == expected_answer.strip().lower():
                correct_count += 1
                feedback.append({
                    'blank_index': i,
                    'student_answer': student_answer,
                    'correct_answer': expected_answer,
                    'is_correct': True
                })
            else:
                feedback.append({
                    'blank_index': i,
                    'student_answer': student_answer or '',
                    'correct_answer': expected_answer,
                    'is_correct': False
                })

        score = (correct_count / total_blanks) * 100 if total_blanks > 0 else 0
        print(f"[WORD_PLACEMENT_DEBUG] Score final: {score}% ({correct_count}/{total_blanks} = {correct_count/total_blanks if total_blanks > 0 else 0})")

    else:
        print(f"[ERROR] Type d'exercice non pris en charge: {exercise.exercise_type}")
        flash(f'Le type d\'exercice {exercise.exercise_type} n\'est pas pris en charge.', 'error')
        return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
    # Créer une nouvelle tentative
    try:
        # Convertir le score en float pour s'assurer qu'il est bien numérique
        score = float(score)
        app.logger.debug(f'Création de la tentative - Score: {score}')
        
        attempt = ExerciseAttempt(
            student_id=current_user.id,
            exercise_id=exercise_id,
            course_id=course_id,
            score=score,
            answers=json.dumps(answers),
            feedback=json.dumps(feedback)
        )
        
        db.session.add(attempt)
        db.session.commit()
        app.logger.debug('Tentative enregistrée avec succès')
        
        flash(f'Exercice soumis avec succès ! Score : {score:.1f}%', 'success')
        return redirect(url_for('view_exercise', exercise_id=exercise_id, course_id=course_id))
        
    except Exception as e:
        app.logger.error(f'Erreur lors de la création de la tentative: {str(e)}')
        db.session.rollback()
        flash('Une erreur est survenue lors de l\'enregistrement de votre tentative.', 'error')
        return redirect(url_for('view_exercise', exercise_id=exercise_id, course_id=course_id))
    return redirect(url_for('view_exercise', exercise_id=exercise_id))

@app.route('/debug/exercises')
@login_required
def debug_exercises():
    if not current_user.is_teacher:
        return "Accès non autorisé", 403
        
    exercises = Exercise.query.all()
    debug_info = []
    
    for ex in exercises:
        debug_info.append({
            'id': ex.id,
            'title': ex.title,
            'type': ex.exercise_type,
            'content': ex.content,
            'parsed_content': ex.get_content()
        })
    
    return render_template('debug_exercises.html', exercises=debug_info)

@app.route('/debug/images')
def debug_images():
    # Lister tous les fichiers dans le dossier uploads
    files = []
    upload_dir = app.config['UPLOAD_FOLDER']
    for filename in os.listdir(upload_dir):
        if filename.startswith('pair_left_'):
            file_path = os.path.join(upload_dir, filename)
            if os.path.isfile(file_path):
                files.append({
                    'name': filename,
                    'url': url_for('static', filename=f'uploads/{filename}'),
                    'size': os.path.getsize(file_path)
                })
    
    return render_template('debug_images.html', files=files)

import random

def generate_word_search_grid(words, max_attempts=3):
    """Génère une grille de mots mêlés à partir d'une liste de mots."""
    if not words:
        return None
        
    # Normaliser les mots (majuscules, pas d'espaces)
    words = [word.strip().upper() for word in words]
    
    # Vérifier la validité des mots
    if any(not word.isalpha() for word in words):
        raise ValueError("Les mots ne doivent contenir que des lettres")
    
    # Trouver la taille de la grille nécessaire
    max_length = max(len(word) for word in words)
    grid_size = max(15, max_length + 2)  # Au moins 15x15 ou assez grand pour le plus long mot
    
    # Directions possibles pour placer les mots
    directions = [
        (0, 1),   # horizontal
        (1, 0),   # vertical
        (1, 1),   # diagonal bas-droite
        (-1, 1),  # diagonal haut-droite
    ]
    
    def can_place_word(grid, word, start_x, start_y, dx, dy):
        """Vérifie si un mot peut être placé à partir d'une position donnée."""
        for i, letter in enumerate(word):
            x = start_x + i * dx
            y = start_y + i * dy
            if not (0 <= x < grid_size and 0 <= y < grid_size):
                return False
            if grid[y][x] and grid[y][x] != letter:
                return False
        return True
    
    def place_word(grid, word):
        """Tente de placer un mot dans la grille."""
        attempts = 100  # Nombre maximum d'essais par mot
        while attempts > 0:
            dx, dy = random.choice(directions)
            if dx == 0:  # horizontal
                x = random.randint(0, grid_size - len(word))
                y = random.randint(0, grid_size - 1)
            elif dy == 0:  # vertical
                x = random.randint(0, grid_size - 1)
                y = random.randint(0, grid_size - len(word))
            else:  # diagonal
                x = random.randint(0, grid_size - len(word))
                y = random.randint(0, grid_size - len(word))
            
            if can_place_word(grid, word, x, y, dx, dy):
                for i, letter in enumerate(word):
                    grid[y + i * dy][x + i * dx] = letter
                return True
            attempts -= 1
        return False
    
    # Essayer de générer une grille valide
    attempt_count = 0
    while attempt_count < max_attempts:
        try:
            # Créer une grille vide
            grid = [['' for _ in range(grid_size)] for _ in range(grid_size)]
            
            # Placer chaque mot
            random.shuffle(words)  # Mélanger les mots pour varier leur placement
            success = True
            for word in words:
                if not place_word(grid, word):
                    success = False
                    break
            
            if success:
                # Remplir les cases vides avec des lettres aléatoires
                letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                for y in range(grid_size):
                    for x in range(grid_size):
                        if not grid[y][x]:
                            grid[y][x] = random.choice(letters)
                return grid
        except Exception:
            pass
            
        attempt_count += 1
    
    return None  # Si on n'a pas réussi à générer une grille valide

@app.route('/exercise/<int:exercise_id>/attempt/<int:attempt_id>')
@login_required
def view_attempt(exercise_id, attempt_id):
    exercise = Exercise.query.get_or_404(exercise_id)
    attempt = ExerciseAttempt.query.get_or_404(attempt_id)
    course_id = request.args.get('course_id', type=int)
    course = Course.query.get(course_id) if course_id else None
    
    # Vérifier que la tentative appartient à l'exercice
    if attempt.exercise_id != exercise_id:
        flash('Cette tentative ne correspond pas à cet exercice.', 'error')
        return redirect(url_for('view_exercise', exercise_id=exercise_id))
    
    # Si c'est un professeur
    if current_user.is_teacher:
        # Vérifier que l'exercice appartient au professeur
        if exercise.teacher_id != current_user.id:
            flash('Vous n\'êtes pas autorisé à voir cette tentative.', 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id))
    # Si c'est un élève
    else:
        # Vérifier que la tentative appartient bien à l'élève
        if attempt.student_id != current_user.id:
            flash('Vous n\'êtes pas autorisé à voir cette tentative.', 'error')
            return redirect(url_for('view_exercise', exercise_id=exercise_id))
    
    # Récupérer le contenu de l'exercice
    content = exercise.get_content()
    if not content:
        flash('Impossible de récupérer le contenu de l\'exercice.', 'error')
        return redirect(url_for('view_exercise', exercise_id=exercise_id))
    
    return render_template('view_attempt.html', 
                         exercise=exercise, 
                         attempt=attempt,
                         course=course,
                         content=content)



# Delete route moved to exercise blueprint in modified_submit.py
    return redirect(url_for('exercise_library'))

# Route conflictuelle supprimée - utiliser handle_exercise_answer dans exercise_routes.py

@app.route('/exercise/preview/<int:exercise_id>')
@login_required
def preview_exercise(exercise_id):
    exercise = Exercise.query.get_or_404(exercise_id)
    content = exercise.get_content()
    
    # Initialiser le contenu par défaut si nécessaire
    if not content:
        content = {}
    
    # Initialiser le contenu en fonction du type d'exercice
    if exercise.exercise_type == 'underline_words':
        if 'sentences' not in content:
            content['sentences'] = []
        # Convertir les phrases en format adapté au template
        for i, sentence in enumerate(content['sentences']):
            if isinstance(sentence, str):
                content['sentences'][i] = {
                    'text': sentence,
                    'words': sentence.split(),
                    'words_to_underline': []
                }
    elif exercise.exercise_type == 'pairs':
        if 'left_items' not in content:
            content['left_items'] = []
        if 'right_items' not in content:
            content['right_items'] = []
        if 'correct_pairs' not in content:
            content['correct_pairs'] = []
    elif exercise.exercise_type == 'qcm':
        if 'questions' not in content:
            content['questions'] = []
        # S'assurer que chaque question a la bonne structure
        for i, question in enumerate(content.get('questions', [])):
            if isinstance(question, dict):
                if 'text' not in question:
                    question['text'] = ''
                if 'choices' not in question:
                    question['choices'] = []
                if 'correct_answer' not in question:
                    question['correct_answer'] = 0
    
    # Choisir le template en fonction du type d'exercice
    template = f'exercise_types/{exercise.exercise_type}.html'
    
    return render_template(template, exercise=exercise, content=content)






# Route supprimée car dupliquée avec /exercise/create

# Route pour la soumission des exercices (intégrée directement)
@app.route('/test-submit')
def test_submit():
    return render_template('test_submit.html')

@app.route('/test-minimal')
def test_minimal():
    """Route pour tester un formulaire ultra-minimal"""
    return render_template('test_minimal_form.html')

@app.route('/test-pure-html')
def test_pure_html():
    """Route pour tester un formulaire HTML pur sans JavaScript"""
    return render_template('test_pure_html.html')

@app.route('/exercise/<int:exercise_id>/answer', methods=['POST'])
# @login_required  # TEMPORAIREMENT DÉSACTIVÉ POUR TEST
def handle_exercise_answer(exercise_id):
    print(f"[ROUTE_DEBUG] Route /exercise/{exercise_id}/answer accessed")
    print(f"[DIAGNOSTIC] handle_exercise_answer called with exercise_id={exercise_id}")
    app.logger.info(f"[DIAGNOSTIC] Function entry - exercise_id={exercise_id}")
    try:
        app.logger.info(f"[SUBMIT_DEBUG] Starting handle_exercise_answer for exercise {exercise_id}")
        exercise = Exercise.query.get_or_404(exercise_id)
        app.logger.info(f"[SUBMIT_DEBUG] Exercise found: {exercise.title}, type: {exercise.exercise_type}")
        course_id = request.form.get('course_id')
        course = Course.query.get(course_id) if course_id else None
        app.logger.info(f"[SUBMIT_DEBUG] Form data keys: {list(request.form.keys())}")
        
        # Vérifier que l'étudiant a accès à l'exercice si c'est via un cours
        # TEMPORAIREMENT DÉSACTIVÉ POUR TEST SANS AUTHENTIFICATION
        # if course and current_user.role == 'student':
        #     class_obj = Class.query.get(course.class_id)
        #     if not class_obj or current_user not in class_obj.students:
        #         flash('Vous n\'avez pas accès à cet exercice.', 'error')
        #         return redirect(url_for('index'))
        
        # Récupérer les réponses et calculer le score
        answers = {}
        score = 0
        max_score = 0
        feedback = {}
        
        if exercise.exercise_type == 'drag_and_drop':
            # Gestion des exercices glisser-déposer
            content = json.loads(exercise.content)
            app.logger.info(f"[DRAG_DROP_DEBUG] Processing drag_and_drop exercise {exercise_id}")
            app.logger.info(f"[DRAG_DROP_DEBUG] Form data: {dict(request.form)}")
            
            user_order = []
            feedback = []
            
            # Récupérer les réponses du formulaire
            for i in range(len(content['draggable_items'])):
                val = request.form.get(f'answer_{i}')
                app.logger.info(f"[DRAG_DROP_DEBUG] answer_{i} = {val} (type: {type(val)})")
                try:
                    idx = int(val) if val is not None else -1
                    user_order.append(idx)
                except (ValueError, TypeError):
                    user_order.append(-1)
            
            correct_order = content.get('correct_order', [])
            app.logger.info(f"[DRAG_DROP_DEBUG] User order: {user_order}")
            app.logger.info(f"[DRAG_DROP_DEBUG] Correct order: {correct_order}")
            
            # Calculer le score
            score_count = 0
            for i, (user_idx, correct_idx) in enumerate(zip(user_order, correct_order)):
                user_item = content['draggable_items'][user_idx] if 0 <= user_idx < len(content['draggable_items']) else "Vide"
                correct_item = content['draggable_items'][correct_idx] if 0 <= correct_idx < len(content['draggable_items']) else "Vide"
                is_correct = (user_idx == correct_idx) and (user_idx != -1)
                
                feedback.append({
                    'zone': i+1,
                    'expected': correct_item,
                    'given': user_item,
                    'is_correct': is_correct
                })
                
                if is_correct:
                    score_count += 1
                    
                app.logger.info(f"[DRAG_DROP_DEBUG] Zone {i+1}: user={user_idx}, correct={correct_idx}, match={is_correct}")
            
            max_score = len(correct_order)
            score = round((score_count / max_score) * 100) if max_score > 0 else 0
            
            # DEBUG: Log final scoring results
            app.logger.info(f"[DRAG_DROP_DEBUG] Score count: {score_count}")
            app.logger.info(f"[DRAG_DROP_DEBUG] Max score: {max_score}")
            app.logger.info(f"[DRAG_DROP_DEBUG] Final score: {score}%")
            
            # Créer le feedback summary pour drag_and_drop
            feedback_summary = {
                'score': score,
                'score_count': score_count,
                'max_score': max_score,
                'details': feedback
            }
            
            # Pour drag_and_drop, on utilise le score en pourcentage mais on sauvegarde aussi le score_count
            answers = {f'answer_{i}': str(user_order[i]) for i in range(len(user_order))}
        
        elif exercise.exercise_type == 'word_search':
            # Gestion des exercices mots mêlés
            content = json.loads(exercise.content)
            app.logger.info(f"[WORD_SEARCH_DEBUG] Processing word_search exercise {exercise_id}")
            app.logger.info(f"[WORD_SEARCH_DEBUG] Form data: {dict(request.form)}")
            
            # Récupérer les mots à trouver
            words_to_find = content.get('words', [])
            if not words_to_find:
                app.logger.error(f"[WORD_SEARCH_DEBUG] No words found in exercise content")
                flash('Erreur: aucun mot trouvé dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[WORD_SEARCH_DEBUG] Words to find: {words_to_find}")
            
            # Récupérer les mots trouvés par l'utilisateur depuis les champs word_0, word_1, etc.
            found_words = []
            for key, value in request.form.items():
                if key.startswith('word_') and value:
                    word = value.strip().upper()
                    if word and word != 'UNDEFINED':  # Éviter les valeurs vides ou non définies
                        found_words.append(word)
            
            # Alternative: récupérer depuis un champ caché ou une liste (pour compatibilité)
            if not found_words:
                found_words_str = request.form.get('found_words', '')
                if found_words_str:
                    found_words = [word.strip().upper() for word in found_words_str.split(',') if word.strip()]
            
            app.logger.info(f"[WORD_SEARCH_DEBUG] Found words by user: {found_words}")
            
            # Calculer le score
            correct_words = []
            incorrect_words = []
            
            for word in found_words:
                if word in words_to_find:
                    correct_words.append(word)
                else:
                    incorrect_words.append(word)
            
            # Mots manqués
            missed_words = [word for word in words_to_find if word not in found_words]
            
            score_count = len(correct_words)
            max_score = len(words_to_find)
            score = (score_count / max_score) * 100 if max_score > 0 else 0
            
            app.logger.info(f"[WORD_SEARCH_DEBUG] Correct words: {correct_words}")
            app.logger.info(f"[WORD_SEARCH_DEBUG] Incorrect words: {incorrect_words}")
            app.logger.info(f"[WORD_SEARCH_DEBUG] Missed words: {missed_words}")
            app.logger.info(f"[WORD_SEARCH_DEBUG] Score: {score_count}/{max_score} = {score}%")
            
            # Créer le feedback
            feedback = {
                'correct_words': correct_words,
                'incorrect_words': incorrect_words,
                'missed_words': missed_words,
                'total_found': len(found_words),
                'total_correct': score_count,
                'total_words': max_score
            }
            
            feedback_summary = {
                'score': score,
                'score_count': score_count,
                'max_score': max_score,
                'details': feedback
            }
            
            # Sauvegarder les réponses
            answers = {'found_words': ','.join(found_words)}
        
        elif exercise.exercise_type == 'pairs':
            # Gestion des exercices d'association de paires
            content = json.loads(exercise.content)
            app.logger.info(f"[PAIRS_DEBUG] Processing pairs exercise {exercise_id}")
            app.logger.info(f"[PAIRS_DEBUG] Form data: {dict(request.form)}")
            
            # Récupérer les paires originales
            original_pairs = content.get('pairs', [])
            if not original_pairs:
                app.logger.error(f"[PAIRS_DEBUG] No pairs found in exercise content")
                flash('Erreur: aucune paire trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            # Créer un dictionnaire des bonnes réponses (left_index -> right_index)
            correct_answers = {}
            for i, pair in enumerate(original_pairs):
                correct_answers[i] = i  # L'index de gauche correspond à l'index de droite original
            
            app.logger.info(f"[PAIRS_DEBUG] Correct answers: {correct_answers}")
            
            # Récupérer les réponses de l'utilisateur
            user_answers = {}
            for i in range(len(original_pairs)):
                left_value = request.form.get(f'left_{i}')
                if left_value:
                    try:
                        right_index = int(left_value)
                        user_answers[i] = right_index
                        app.logger.info(f"[PAIRS_DEBUG] User paired left_{i} with right_{right_index}")
                    except (ValueError, TypeError):
                        app.logger.warning(f"[PAIRS_DEBUG] Invalid value for left_{i}: {left_value}")
            
            app.logger.info(f"[PAIRS_DEBUG] User answers: {user_answers}")
            
            # Calculer le score
            score_count = 0
            feedback = []
            
            for left_index in range(len(original_pairs)):
                user_right_index = user_answers.get(left_index, -1)
                correct_right_index = correct_answers.get(left_index, -1)
                is_correct = user_right_index == correct_right_index
                
                if is_correct:
                    score_count += 1
                
                # Créer le feedback pour cette paire
                left_item = original_pairs[left_index]['left'] if left_index < len(original_pairs) else None
                expected_right = original_pairs[correct_right_index]['right'] if 0 <= correct_right_index < len(original_pairs) else None
                given_right = original_pairs[user_right_index]['right'] if 0 <= user_right_index < len(original_pairs) else None
                
                feedback.append({
                    'left_index': left_index,
                    'left_item': left_item,
                    'expected_right': expected_right,
                    'given_right': given_right,
                    'is_correct': is_correct
                })
                
                app.logger.info(f"[PAIRS_DEBUG] Pair {left_index}: user={user_right_index}, correct={correct_right_index}, is_correct={is_correct}")
            
            max_score = len(original_pairs)
            score = (score_count / max_score) * 100 if max_score > 0 else 0
            
            # DEBUG: Log final scoring results
            app.logger.info(f"[PAIRS_DEBUG] Score count: {score_count}")
            app.logger.info(f"[PAIRS_DEBUG] Max score: {max_score}")
            app.logger.info(f"[PAIRS_DEBUG] Final score: {score}%")
            
            # Créer le feedback summary pour pairs
            feedback_summary = {
                'score': score,
                'score_count': score_count,
                'max_score': max_score,
                'details': feedback
            }
            
            # Pour pairs, on utilise le score en pourcentage
            answers = {f'left_{i}': str(user_answers.get(i, -1)) for i in range(len(original_pairs))}
        
        elif exercise.exercise_type == 'underline_words':
            # Gestion des exercices "Souligner les mots"
            content = json.loads(exercise.content)
            app.logger.info(f"[UNDERLINE_DEBUG] Processing underline_words exercise {exercise_id}")
            app.logger.info(f"[UNDERLINE_DEBUG] Form data: {dict(request.form)}")
            
            # Récupérer les phrases originales (structure 'words' au lieu de 'sentences')
            sentences = content.get('words', [])  # CORRECTION: utiliser 'words' au lieu de 'sentences'
            if not sentences:
                app.logger.error(f"[UNDERLINE_DEBUG] No words/sentences found in exercise content")
                app.logger.error(f"[UNDERLINE_DEBUG] Available content keys: {list(content.keys())}")
                flash('Erreur: aucune phrase trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[UNDERLINE_DEBUG] Found {len(sentences)} sentences to process")
            
            # Calculer le score pour chaque phrase
            total_sentences = len(sentences)
            correct_sentences = 0
            feedback = []
            
            for i, sentence_data in enumerate(sentences):
                # Récupérer les mots que l'utilisateur a soulignés pour cette phrase
                # Le template utilise selected_words_{i} avec des mots séparés par des virgules
                user_underlined_str = request.form.get(f'selected_words_{i}', '')
                user_underlined = [word.strip() for word in user_underlined_str.split(',') if word.strip()]
                app.logger.info(f"[UNDERLINE_DEBUG] Sentence {i}: user underlined string '{user_underlined_str}'")
                app.logger.info(f"[UNDERLINE_DEBUG] Sentence {i}: user underlined list {user_underlined}")
                
                # Récupérer les mots qui devaient être soulignés
                expected_words = sentence_data.get('words_to_underline', [])
                app.logger.info(f"[UNDERLINE_DEBUG] Sentence {i}: expected words {expected_words}")
                
                # Normaliser les mots (minuscules, sans ponctuation, sans apostrophes)
                def normalize_word(word):
                    # Convertir en minuscules
                    normalized = word.lower()
                    # Supprimer la ponctuation au début et à la fin
                    normalized = normalized.strip('.,!?;:\'"()[]{}«»')
                    # Gérer les apostrophes (l'autoroute -> autoroute)
                    if normalized.startswith(("l'", "d'", "n'", "m'", "t'", "s'", "c'", "j'", "qu'")):
                        normalized = normalized[2:]  # Supprimer l' d' etc.
                    elif normalized.startswith(("la", "le", "les", "un", "une", "du", "de", "des")):
                        # Garder les articles complets
                        pass
                    return normalized
                
                user_words_normalized = set(normalize_word(word) for word in user_underlined if word.strip())
                expected_words_normalized = set(normalize_word(word) for word in expected_words if word.strip())
                
                app.logger.info(f"[UNDERLINE_DEBUG] Sentence {i}: user normalized {user_words_normalized}")
                app.logger.info(f"[UNDERLINE_DEBUG] Sentence {i}: expected normalized {expected_words_normalized}")
                
                # Vérifier si la phrase est correcte (tous les mots attendus sont soulignés)
                is_correct = user_words_normalized == expected_words_normalized
                if is_correct:
                    correct_sentences += 1
                
                # Créer le feedback pour cette phrase
                sentence_text = sentence_data.get('text', '')  # CORRECTION: utiliser 'text' au lieu de 'words'
                feedback.append({
                    'sentence_index': i,
                    'sentence_text': sentence_text,
                    'user_underlined': list(user_words_normalized),
                    'expected_words': list(expected_words_normalized),
                    'is_correct': is_correct,
                    'missing_words': list(expected_words_normalized - user_words_normalized),
                    'extra_words': list(user_words_normalized - expected_words_normalized)
                })
                
                app.logger.info(f"[UNDERLINE_DEBUG] Sentence {i}: is_correct={is_correct}")
            
            # Calculer le score final
            score = (correct_sentences / total_sentences) * 100 if total_sentences > 0 else 0
            
            # DEBUG: Log final scoring results
            app.logger.info(f"[UNDERLINE_DEBUG] Correct sentences: {correct_sentences}")
            app.logger.info(f"[UNDERLINE_DEBUG] Total sentences: {total_sentences}")
            app.logger.info(f"[UNDERLINE_DEBUG] Final score: {score}%")
            
            # Créer le feedback summary pour underline_words
            feedback_summary = {
                'score': score,
                'correct_sentences': correct_sentences,
                'total_sentences': total_sentences,
                'details': feedback
            }
            
            # Pour underline_words, on utilise le score en pourcentage
            answers = {f'selected_words_{i}': request.form.get(f'selected_words_{i}', '') for i in range(len(sentences))}
        
        elif exercise.exercise_type == 'dictation':
            # Gestion des exercices de dictée
            content = json.loads(exercise.content)
            app.logger.info(f"[DICTATION_DEBUG] Processing dictation exercise {exercise_id}")
            app.logger.info(f"[DICTATION_DEBUG] Form data: {dict(request.form)}")
            
            # Récupérer les phrases de référence
            reference_sentences = content.get('sentences', [])
            if not reference_sentences:
                app.logger.error(f"[DICTATION_DEBUG] No sentences found in exercise content")
                flash('Erreur: aucune phrase trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[DICTATION_DEBUG] Found {len(reference_sentences)} sentences to check")
            
            # Calculer le score pour chaque phrase
            total_sentences = len(reference_sentences)
            correct_sentences = 0
            feedback = []
            user_answers = {}
            
            for i, reference_sentence in enumerate(reference_sentences):
                # Récupérer la réponse de l'utilisateur
                user_answer = request.form.get(f'dictation_answer_{i}', '').strip()
                reference_text = reference_sentence.strip()
                
                user_answers[f'dictation_answer_{i}'] = user_answer
                
                app.logger.info(f"[DICTATION_DEBUG] Sentence {i}: user='{user_answer}' vs reference='{reference_text}'")
                
                # Normaliser les textes pour la comparaison
                def normalize_text(text):
                    # Convertir en minuscules
                    normalized = text.lower()
                    # Supprimer la ponctuation
                    import string
                    normalized = normalized.translate(str.maketrans('', '', string.punctuation))
                    # Supprimer les espaces multiples
                    normalized = ' '.join(normalized.split())
                    return normalized
                
                user_normalized = normalize_text(user_answer)
                reference_normalized = normalize_text(reference_text)
                
                app.logger.info(f"[DICTATION_DEBUG] Sentence {i}: normalized user='{user_normalized}' vs reference='{reference_normalized}'")
                
                # Vérifier si la phrase est correcte
                is_correct = user_normalized == reference_normalized
                if is_correct:
                    correct_sentences += 1
                
                # Calculer la similarité (pourcentage de mots corrects)
                user_words = user_normalized.split()
                reference_words = reference_normalized.split()
                
                if reference_words:
                    # Compter les mots corrects (même position)
                    correct_words = 0
                    max_length = max(len(user_words), len(reference_words))
                    
                    for j in range(max_length):
                        user_word = user_words[j] if j < len(user_words) else ''
                        ref_word = reference_words[j] if j < len(reference_words) else ''
                        if user_word == ref_word and user_word != '':
                            correct_words += 1
                    
                    similarity = (correct_words / len(reference_words)) * 100
                else:
                    similarity = 100 if not user_words else 0
                
                # Créer le feedback pour cette phrase
                feedback.append({
                    'sentence_index': i,
                    'user_answer': user_answer,
                    'reference_sentence': reference_text,
                    'is_correct': is_correct,
                    'similarity': round(similarity, 1),
                    'status': 'Correct' if is_correct else f'Similarité: {round(similarity, 1)}%'
                })
            
            # Calculer le score final
            max_score = total_sentences
            score_count = correct_sentences
            score = round((score_count / max_score) * 100) if max_score > 0 else 0
            
            app.logger.info(f"[DICTATION_DEBUG] Final score: {score_count}/{max_score} = {score}%")
            
            feedback_summary = {
                'score': score,
                'correct_sentences': correct_sentences,
                'total_sentences': total_sentences,
                'details': feedback
            }
            # Pour dictation, on utilise le score en pourcentage
            answers = user_answers
        
        elif exercise.exercise_type == 'image_labeling':
            # Gestion des exercices d'étiquetage d'image
            content = json.loads(exercise.content)
            app.logger.info(f"[IMAGE_LABELING_DEBUG] Processing image_labeling exercise {exercise_id}")
            app.logger.info(f"[IMAGE_LABELING_DEBUG] Form data: {dict(request.form)}")
            
            # Récupérer les zones définies dans l'exercice
            zones = content.get('zones', [])
            if not zones:
                app.logger.error(f"[IMAGE_LABELING_DEBUG] No zones found in exercise content")
                flash('Erreur: aucune zone trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[IMAGE_LABELING_DEBUG] Found {len(zones)} zones to check")
            
            # Récupérer les réponses de l'utilisateur depuis le JSON
            user_answers_json = request.form.get('user_answers', '{}')
            try:
                user_answers_data = json.loads(user_answers_json)
                app.logger.info(f"[IMAGE_LABELING_DEBUG] User answers: {user_answers_data}")
            except json.JSONDecodeError:
                app.logger.error(f"[IMAGE_LABELING_DEBUG] Invalid JSON in user_answers: {user_answers_json}")
                user_answers_data = {}
            
            # Calculer le score
            total_zones = len(zones)
            correct_zones = 0
            feedback = []
            
            for i, zone in enumerate(zones):
                zone_id = str(i + 1)  # Les zones sont numérotées à partir de 1
                expected_label = zone.get('label', '')
                user_label = user_answers_data.get(zone_id, '')
                
                is_correct = user_label.lower().strip() == expected_label.lower().strip()
                if is_correct:
                    correct_zones += 1
                
                feedback.append({
                    'zone_id': zone_id,
                    'expected_label': expected_label,
                    'user_label': user_label,
                    'is_correct': is_correct,
                    'status': 'Correct' if is_correct else f'Attendu: {expected_label}, Réponse: {user_label}'
                })
                
                app.logger.info(f"[IMAGE_LABELING_DEBUG] Zone {zone_id}: expected='{expected_label}', user='{user_label}', correct={is_correct}")
            
            # Calculer le score final
            max_score = total_zones
            score_count = correct_zones
            score = round((score_count / max_score) * 100) if max_score > 0 else 0
            
            app.logger.info(f"[IMAGE_LABELING_DEBUG] Final score: {score_count}/{max_score} = {score}%")
            
            feedback_summary = {
                'score': score,
                'correct_zones': correct_zones,
                'total_zones': total_zones,
                'details': feedback
            }
            
            # Pour image_labeling, on utilise le score en pourcentage
            answers = user_answers_data
        
        elif exercise.exercise_type == 'qcm_multichoix':
            # Gestion des exercices QCM Multichoix
            content = json.loads(exercise.content)
            app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Processing qcm_multichoix exercise {exercise_id}")
            app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Form data: {dict(request.form)}")
            
            # Récupérer les questions de l'exercice
            questions = content.get('questions', [])
            if not questions:
                app.logger.error(f"[QCM_MULTICHOIX_DEBUG] No questions found in exercise content")
                flash('Erreur: aucune question trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Found {len(questions)} questions to check")
            
            # Calculer le score
            total_questions = len(questions)
            correct_questions = 0
            feedback_details = []
            user_answers_data = {}
            
            for question_index, question in enumerate(questions):
                # Récupérer les réponses de l'utilisateur pour cette question
                user_selected = request.form.getlist(f'question_{question_index}[]')
                app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Raw user_selected for question {question_index}: {user_selected}")
                user_selected_indices = [int(idx) for idx in user_selected if idx.isdigit()]
                
                # Récupérer les bonnes réponses pour cette question
                correct_options = question.get('correct_options', [])
                
                app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Question {question_index}: user={user_selected_indices}, correct={correct_options}")
                app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Question {question_index}: user_set={set(user_selected_indices)}, correct_set={set(correct_options)}")
                
                # Vérifier si les réponses correspondent exactement
                is_correct = set(user_selected_indices) == set(correct_options)
                app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Question {question_index}: is_correct={is_correct}")
                if is_correct:
                    correct_questions += 1
                
                # Créer le feedback pour cette question
                user_options = [question['options'][i] for i in user_selected_indices if i < len(question['options'])]
                correct_options_text = [question['options'][i] for i in correct_options if i < len(question['options'])]
                
                feedback_details.append({
                    'question_index': question_index,
                    'question_text': question.get('question', ''),
                    'user_selected': user_options,
                    'correct_options': correct_options_text,
                    'is_correct': is_correct,
                    'status': 'Correct' if is_correct else f'Attendu: {", ".join(correct_options_text)}, Réponse: {", ".join(user_options) if user_options else "Aucune réponse"}'
                })
                
                # Sauvegarder les réponses utilisateur
                user_answers_data[f'question_{question_index}'] = user_selected_indices
            
            # Calculer le score final
            max_score = total_questions
            score_count = correct_questions
            score = round((score_count / max_score) * 100) if max_score > 0 else 0
            
            app.logger.info(f"[QCM_MULTICHOIX_DEBUG] Final score: {score_count}/{max_score} = {score}%")
            
            feedback_summary = {
                'score': score,
                'correct_questions': correct_questions,
                'total_questions': total_questions,
                'details': feedback_details
            }
            
            # Pour qcm_multichoix, on utilise le score en pourcentage
            answers = user_answers_data
        
        elif exercise.exercise_type == 'flashcards':
            # Gestion des exercices Flashcards
            content = json.loads(exercise.content)
            app.logger.info(f"[FLASHCARDS_DEBUG] Processing flashcards exercise {exercise_id}")
            app.logger.info(f"[FLASHCARDS_DEBUG] Form data: {dict(request.form)}")
            
            # Récupérer les cartes de l'exercice
            cards = content.get('cards', [])
            if not cards:
                app.logger.error(f"[FLASHCARDS_DEBUG] No cards found in exercise content")
                flash('Erreur: aucune carte trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[FLASHCARDS_DEBUG] Found {len(cards)} cards to check")
            
            # Récupérer le score et les réponses depuis le formulaire
            final_score = request.form.get('final_score', '0')
            correct_answers_count = request.form.get('correct_answers', '0')
            total_answers_count = request.form.get('total_answers', '0')
            
            try:
                score = float(final_score)
                score_count = int(correct_answers_count)
                max_score = int(total_answers_count)
            except (ValueError, TypeError):
                app.logger.error(f"[FLASHCARDS_DEBUG] Invalid score data: final_score={final_score}, correct={correct_answers_count}, total={total_answers_count}")
                score = 0
                score_count = 0
                max_score = len(cards)
            
            app.logger.info(f"[FLASHCARDS_DEBUG] Score received: {score_count}/{max_score} = {score}%")
            
            # Récupérer les réponses individuelles pour chaque carte
            user_answers = {}
            feedback_details = []
            
            for i in range(len(cards)):
                card_answer = request.form.get(f'card_{i}_answer', '')
                card_correct = request.form.get(f'card_{i}_correct', 'false') == 'true'
                
                user_answers[f'card_{i}'] = {
                    'answer': card_answer,
                    'correct': card_correct,
                    'expected': cards[i].get('answer', '') if i < len(cards) else ''
                }
                
                feedback_details.append({
                    'card_index': i,
                    'question': cards[i].get('question', '') if i < len(cards) else '',
                    'expected_answer': cards[i].get('answer', '') if i < len(cards) else '',
                    'user_answer': card_answer,
                    'is_correct': card_correct
                })
            
            feedback_summary = {
                'score': score,
                'score_count': score_count,
                'max_score': max_score,
                'total_cards': len(cards),
                'details': feedback_details
            }
            
            app.logger.info(f"[FLASHCARDS_DEBUG] Final feedback: {feedback_summary}")
            
            # Pour flashcards, on utilise le score calculé côté frontend
            answers = user_answers
        
        elif exercise.exercise_type == 'fill_in_blanks':
            # Gestion des exercices Texte à trous avec la même logique que Mots à placer
            content = json.loads(exercise.content)
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Processing fill_in_blanks exercise {exercise_id}")
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Form data: {dict(request.form)}")
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Exercise content keys: {list(content.keys())}")
            
            # Compter le nombre réel de blancs dans le contenu
            total_blanks_in_content = 0
            
            # Analyser le format de l'exercice et compter les blancs réels
            # CORRECTION: Éviter le double comptage entre 'text' et 'sentences'
            # Priorité à 'sentences' s'il existe, sinon utiliser 'text'
            if 'sentences' in content:
                sentences_blanks = sum(s.count('___') for s in content['sentences'])
                total_blanks_in_content = sentences_blanks
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Format 'sentences' détecté: {sentences_blanks} blancs dans sentences")
                # Log détaillé pour chaque phrase et ses blancs
                for i, sentence in enumerate(content['sentences']):
                    blanks_in_sentence = sentence.count('___')
                    app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Phrase {i}: '{sentence}' contient {blanks_in_sentence} blancs")
            elif 'text' in content:
                text_blanks = content['text'].count('___')
                total_blanks_in_content = text_blanks
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Format 'text' détecté: {text_blanks} blancs dans text")
            
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Total blancs trouvés dans le contenu: {total_blanks_in_content}")
            # Log détaillé pour chaque phrase et ses blancs
            if 'sentences' in content:
                for i, sentence in enumerate(content['sentences']):
                    blanks_in_sentence = sentence.count('___')
                    app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Phrase {i}: '{sentence}' contient {blanks_in_sentence} blancs")
            
            # Récupérer les réponses correctes (peut être 'words' ou 'available_words')
            correct_answers = content.get('words', [])
            if not correct_answers:
                correct_answers = content.get('available_words', [])
            
            if not correct_answers:
                app.logger.error(f"[FILL_IN_BLANKS_DEBUG] No correct answers found in exercise content")
                flash('Erreur: aucune réponse correcte trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Found {len(correct_answers)} correct answers: {correct_answers}")
            
            # Utiliser le nombre réel de blancs trouvés dans le contenu
            total_blanks = max(total_blanks_in_content, len(correct_answers))
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Using total_blanks = {total_blanks}")
            
            correct_blanks = 0
            feedback_details = []
            user_answers_data = {}
            
            # Vérifier si c'est un exercice de rangement par ordre
            is_ordering, ordering_type = is_ordering_exercise(exercise.description)
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Détection exercice de rangement: {is_ordering}, type: {ordering_type}")

            # Si c'est un exercice de rangement, utiliser la logique spécifique
            if is_ordering:
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Traitement d'un exercice de rangement {ordering_type}")
                
                # Récupérer toutes les réponses de l'utilisateur
                user_answers_list = []
                for i in range(total_blanks):
                    user_answer = request.form.get(f'answer_{i}', '').strip()
                    user_answers_list.append(user_answer)
                    user_answers_data[f'answer_{i}'] = user_answer
                
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Réponses utilisateur: {user_answers_list}")
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Réponses correctes: {correct_answers}")
                
                # Évaluer les réponses avec la logique de rangement
                correct_count, feedback_list = evaluate_ordering_exercise(user_answers_list, correct_answers, ordering_type)
                correct_blanks = correct_count
                
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Résultat évaluation: {correct_count}/{total_blanks} corrects")
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Feedback détaillé: {feedback_list}")
                
                # Créer le feedback détaillé
                for i in range(len(user_answers_list)):
                    user_answer = user_answers_list[i] if i < len(user_answers_list) else ''
                    correct_answer = correct_answers[i] if i < len(correct_answers) else ''
                    is_correct = feedback_list[i] if i < len(feedback_list) else False
                    
                    # Déterminer l'index de la phrase à laquelle appartient ce blanc
                    sentence_index = -1
                    local_blank_index = -1
                    if 'sentences' in content:
                        sentence_index, local_blank_index = get_blank_location(i, content['sentences'])
                    
                    feedback_details.append({
                        'blank_index': i,
                        'user_answer': user_answer or '',
                        'correct_answer': correct_answer,
                        'is_correct': is_correct,
                        'status': 'Correct' if is_correct else f'Attendu: {correct_answer}, Réponse: {user_answer or "Vide"}',
                        'sentence_index': sentence_index,
                        'sentence': content['sentences'][sentence_index] if sentence_index >= 0 and 'sentences' in content else ''
                    })
            else:
                # Vérifier chaque blanc individuellement - Même logique que word_placement
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Traitement de {total_blanks} blancs au total (exercice standard)")
                for i in range(total_blanks):
                    # Récupérer la réponse de l'utilisateur pour ce blanc
                    user_answer = request.form.get(f'answer_{i}', '').strip()
                    
                    # Récupérer la réponse correcte correspondante
                    correct_answer = correct_answers[i] if i < len(correct_answers) else ''
                    
                    app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Blank {i}:")
                    app.logger.info(f"  - Réponse étudiant (answer_{i}): {user_answer}")
                    app.logger.info(f"  - Réponse attendue: {correct_answer}")
                    
                    # Vérifier si la réponse est correcte (insensible à la casse)
                    is_correct = user_answer and user_answer.strip().lower() == correct_answer.strip().lower()
                    if is_correct:
                        correct_blanks += 1
                    
                    # Créer le feedback pour ce blanc
                    # Déterminer l'index de la phrase à laquelle appartient ce blanc
                    sentence_index = -1
                    local_blank_index = -1
                    if 'sentences' in content:
                        sentence_index, local_blank_index = get_blank_location(i, content['sentences'])
                        app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Blank {i} est dans la phrase {sentence_index}, position locale {local_blank_index}")
                    
                    feedback_details.append({
                        'blank_index': i,
                        'user_answer': user_answer or '',
                        'correct_answer': correct_answer,
                        'is_correct': is_correct,
                        'status': 'Correct' if is_correct else f'Attendu: {correct_answer}, Réponse: {user_answer or "Vide"}',
                        'sentence_index': sentence_index,
                        'sentence': content['sentences'][sentence_index] if sentence_index >= 0 and 'sentences' in content else ''
                    })
                    
                    # Sauvegarder les réponses utilisateur
                    user_answers_data[f'answer_{i}'] = user_answer
            # Calculer le score final basé sur le nombre réel de blancs - Exactement comme word_placement
            score = (correct_blanks / total_blanks) * 100 if total_blanks > 0 else 0
            
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Score final: {score}% ({correct_blanks}/{total_blanks})")
            
            feedback_summary = {
                'score': score,
                'correct_blanks': correct_blanks,
                'total_blanks': total_blanks,
                'details': feedback_details
            }
            
            # Pour fill_in_blanks, on utilise le score calculé
            answers = user_answers_data
        
                # Logique de scoring pour fill_in_blanks
        elif exercise.exercise_type == 'fill_in_blanks':
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Traitement de l'exercice fill_in_blanks ID {exercise_id}")
            content = exercise.get_content()
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Contenu de l'exercice: {content}")
            
            total_blanks_in_content = 0
            
            # Analyser le format de l'exercice et compter les blancs réels
            # CORRECTION: Éviter le double comptage entre 'text' et 'sentences'
            # Priorité à 'sentences' s'il existe, sinon utiliser 'text'
            if 'sentences' in content:
                sentences_blanks = sum(s.count('___') for s in content['sentences'])
                total_blanks_in_content = sentences_blanks
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Format 'sentences' détecté: {sentences_blanks} blancs dans sentences")
                # Log détaillé pour chaque phrase et ses blancs
                for i, sentence in enumerate(content['sentences']):
                    blanks_in_sentence = sentence.count('___')
                    app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Phrase {i}: '{sentence}' contient {blanks_in_sentence} blancs")
            elif 'text' in content:
                text_blanks = content['text'].count('___')
                total_blanks_in_content = text_blanks
                app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Format 'text' détecté: {text_blanks} blancs dans text")
            
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Total blancs trouvés dans le contenu: {total_blanks_in_content}")
            # Log détaillé pour chaque phrase et ses blancs
            if 'sentences' in content:
                for i, sentence in enumerate(content['sentences']):
                    blanks_in_sentence = sentence.count('___')
                    app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Phrase {i}: '{sentence}' contient {blanks_in_sentence} blancs")
            
            # Récupérer les réponses correctes (peut être 'words' ou 'available_words')
            correct_answers = content.get('words', [])
            if not correct_answers:
                correct_answers = content.get('available_words', [])
            
            if not correct_answers:
                app.logger.error(f"[FILL_IN_BLANKS_DEBUG] No correct answers found in exercise content")
                flash('Erreur: aucune réponse correcte trouvée dans l\'exercice.', 'error')
                return redirect(url_for('view_exercise', exercise_id=exercise_id))
            
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Found {len(correct_answers)} correct answers: {correct_answers}")
            
            # Utiliser le nombre réel de blancs trouvés dans le contenu
            total_blanks = max(total_blanks_in_content, len(correct_answers))
            app.logger.info(f"[FILL_IN_BLANKS_DEBUG] Using total_blanks = {total_blanks}")
            
            correct_blanks = 0
            feedback_details = []
            user_answers_data = {}
            
    except Exception as e:
        app.logger.error(f"[ERROR] Une erreur s'est produite: {str(e)}")
        flash('Une erreur s\'est produite lors du traitement de votre réponse.', 'error')
        return redirect(url_for('view_exercise', exercise_id=exercise_id))
