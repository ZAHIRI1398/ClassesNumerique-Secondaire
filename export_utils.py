from flask import send_file
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import os
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_class_excel(class_data):
    """Génère un fichier Excel avec les statistiques de la classe en utilisant directement openpyxl"""
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Limiter le nom de la feuille à 31 caractères (limite Excel)
    sheet_name = f"Classe {class_data['class'].name}"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    ws.title = sheet_name
    
    # Définir les en-têtes
    headers = ['Nom', 'Exercices complétés', 'Total exercices', 'Score moyen (%)', 'Progression (%)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Ajouter les données des élèves
    for row_idx, student in enumerate(class_data['students'], 2):
        # Nom
        ws.cell(row=row_idx, column=1).value = student['student'].name or student['student'].username
        
        # Exercices complétés
        ws.cell(row=row_idx, column=2).value = student['completed_exercises']
        
        # Total exercices
        ws.cell(row=row_idx, column=3).value = student['total_exercises']
        
        # Score moyen
        score_cell = ws.cell(row=row_idx, column=4)
        if student['average_score'] is not None:
            score_cell.value = f"{student['average_score']:.1f}"
        else:
            score_cell.value = "N/A"
        
        # Progression
        progress_cell = ws.cell(row=row_idx, column=5)
        if student['total_exercises'] > 0:
            progress = (student['completed_exercises'] / student['total_exercises']) * 100
            progress_cell.value = f"{progress:.1f}"
        else:
            progress_cell.value = "0.0"
    
    # Ajouter des informations générales
    ws.cell(row=1, column=7).value = "Informations générales"
    ws.cell(row=1, column=7).font = Font(bold=True)
    
    ws.cell(row=2, column=7).value = f"Nombre d'élèves: {len(class_data['students'])}"
    ws.cell(row=3, column=7).value = f"Total exercices: {class_data['total_exercises']}"
    ws.cell(row=4, column=7).value = f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Ajuster la largeur des colonnes
    for col_idx in range(1, 6):
        max_length = 0
        column = get_column_letter(col_idx)
        
        for row_idx in range(1, len(class_data['students']) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    # Créer un buffer pour stocker le fichier Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def generate_class_pdf(class_data):
    """Génère un fichier PDF avec les statistiques de la classe"""
    # Créer un buffer pour stocker le PDF
    buffer = io.BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Titre
    elements.append(Paragraph(f"Statistiques de la classe: {class_data['class'].name}", title_style))
    elements.append(Spacer(1, 12))
    
    # Informations générales
    elements.append(Paragraph("Informations générales", subtitle_style))
    elements.append(Spacer(1, 6))
    info_data = [
        ["Nombre d'élèves", str(len(class_data['students']))],
        ["Total exercices", str(class_data['total_exercises'])]
    ]
    info_table = Table(info_data, colWidths=[200, 100])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))
    
    # Tableau des élèves
    elements.append(Paragraph("Résultats par élève", subtitle_style))
    elements.append(Spacer(1, 6))
    
    # En-tête du tableau
    student_data = [["Élève", "Exercices complétés", "Score moyen (%)", "Progression (%)"]]
    
    # Données des élèves
    for student in class_data['students']:
        name = student['student'].name or student['student'].username
        completed = f"{student['completed_exercises']} / {student['total_exercises']}"
        avg_score = f"{student['average_score']:.1f}" if student['average_score'] is not None else "N/A"
        progress = f"{(student['completed_exercises'] / student['total_exercises'] * 100):.1f}" if student['total_exercises'] > 0 else "0.0"
        
        student_data.append([name, completed, avg_score, progress])
    
    # Créer le tableau
    student_table = Table(student_data, colWidths=[150, 100, 100, 100])
    student_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
    ]))
    elements.append(student_table)
    
    # Construire le PDF
    doc.build(elements)
    
    # Réinitialiser le pointeur du buffer
    buffer.seek(0)
    
    return buffer
